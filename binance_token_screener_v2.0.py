#!/usr/bin/env python3
"""
Binance Token Screener v2.0 - Production Release
币安代币筛选器 v2.0 - 生产版本

A comprehensive cryptocurrency analysis system that fetches real-time data from Binance APIs,
performs multi-dimensional analysis, and generates comprehensive reports with Google Sheets integration.

Core Features:
- OAuth authentication for Google Sheets integration
- Real-time data from Binance Spot and Futures markets
- CoinGecko integration for accurate market cap data
- 8 specialized analysis worksheets
- Automated daily scheduling system
- Comprehensive error handling and logging

Version 2.0 (July 29, 2025):
- Fixed 低量高市值 tab creation and data display
- Enhanced 14-day historical data calculations
- Improved column configurations for all tabs
- Updated to analyze top 150 market cap tokens
- Production-ready with full bug fixes

Author: Binance Token Analysis Team
Version: 2.0
Date: 2025-07-29
"""

import pandas as pd
from datetime import datetime, timedelta
import os
import sqlite3
import json
import time
import requests
import argparse
import sys
# from final_fixed_screener import FinalFixedScreener  # 已移除
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
import gspread
import shutil
# 网络重试机制将使用简单的time.sleep
# Excel处理将使用pandas内置功能
from coingecko_integration import CoinGeckoClient
from data_supplement import DataSupplementer

class BinanceTokenScreenerV1:
    """
    Binance Token Screener v2.0 - Production Release
    OAuth-based token analysis system with CoinGecko integration
    """
    
    def __init__(self, auto_mode=False):
        # Version information
        self.version = "1.1"
        self.release_date = "2025-07-28"
        
        # Auto mode flag for automated runs
        self.auto_mode = auto_mode
        
        # OAuth configuration
        self.SCOPES = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        self.oauth_creds_file = 'oauth_credentials.json'
        self.token_file = 'token.json'
        self.creds = None
        self.gc = None
        
        # Core components - 功能已集成到主类中

        # 创建输出文件夹
        self.output_folder = self.create_output_folder()

        # Default settings (configurable via user interaction) - Updated per 7.15 requirements
        self.spot_count = 80
        self.futures_count = 80
        self.user_email = "seanxx.eth@gmail.com"
        
        # Google Sheets configuration - Updated per 7.15 requirements
        self.base_sheet_name = "Binance_Token_screener_Analysis"
        self.sheet_name = self.generate_daily_sheet_name()
        self.worksheets_config = {
            'raw_data': '原始数据',
            'volume_ranking': '交易量排行',
            'market_cap_ranking': '市值排行',
            'gainers_losers': '涨跌排行',
            'futures_focus': '期货专注',
            'daily_gainers_history': '每日涨幅榜',
            'low_volume_high_mcap': '低量高市值',
            'high_volume_low_mcap': '高市值低期货量',
            'low_mcap_high_volume': '低市值高交易量'
        }

        # Excluded tokens per 7.15 requirements
        self.excluded_tokens = {'ALPACA', 'BNX', 'LINA', 'VIDT', 'AGIX', 'FTM'}

        # Stablecoins to exclude from ALL tabs (7.28 requirement 2)
        self.stablecoins_to_exclude = {'USDC', 'FDUSD', 'USD1'}

        # Major coins to ignore in anomaly detection only (not all tabs)
        self.major_coins_to_ignore = {'BTC', 'ETH', 'SOL', 'XRP', 'BNB'}
        
        # Initialize CoinGecko client for real market cap data
        self.coingecko_client = CoinGeckoClient()
        print("🦎 CoinGecko客户端已初始化")
        
        # Initialize data supplementer for missing tokens
        self.data_supplementer = DataSupplementer()
        print("📊 数据补全器已初始化")

    def apply_selective_exclusions(self, df, tab_name):
        """
        应用选择性排除逻辑 - 7.28需求2
        稳定币从ALL标签页排除，主要币种仅从分析标签页排除
        """
        original_count = len(df)
        
        # 检查是否有base_asset列 (每日涨幅榜有特殊结构)
        if 'base_asset' not in df.columns:
            # 对于每日涨幅榜，应用特殊的排除逻辑
            if tab_name == '每日涨幅榜':
                return self.apply_exclusions_to_daily_gainers(df, tab_name)
            else:
                print(f"  📊 {tab_name}: 特殊格式数据，跳过排除逻辑")
                return df
        
        # 排除稳定币 (所有标签页)
        if not df.empty:
            df = df[~df['base_asset'].isin(self.stablecoins_to_exclude)].copy()
            stablecoin_excluded = original_count - len(df)
            if stablecoin_excluded > 0:
                print(f"  📊 {tab_name}: 排除 {stablecoin_excluded} 个稳定币")
        
        # 排除主要币种 (仅分析标签页，不包括核心数据标签页)
        # 更新：期货专注和低市值高交易量也要排除主要币种
        # 低量高市值不应该排除主要币种，因为它要显示市值前120中期货量最低的
        core_tabs = {'原始数据', '交易量排行', '低量高市值'}
        analysis_tabs = {'市值排行', '涨跌排行', '每日涨幅榜', '高市值低期货量', '低市值高交易量', '期货专注'}
        
        if tab_name in analysis_tabs and not df.empty:
            before_major_exclusion = len(df)
            df = df[~df['base_asset'].isin(self.major_coins_to_ignore)].copy()
            major_excluded = before_major_exclusion - len(df)
            if major_excluded > 0:
                print(f"  📊 {tab_name}: 排除 {major_excluded} 个主要币种 (分析标签页)")
        
        # 排除废弃代币 (所有标签页)
        if not df.empty:
            before_deprecated = len(df)
            df = df[~df['base_asset'].isin(self.excluded_tokens)].copy()
            deprecated_excluded = before_deprecated - len(df)
            if deprecated_excluded > 0:
                print(f"  📊 {tab_name}: 排除 {deprecated_excluded} 个废弃代币")
        
        total_excluded = original_count - len(df)
        if total_excluded > 0:
            print(f"  ✅ {tab_name}: 总共排除 {total_excluded} 个代币，保留 {len(df)} 个")
        
        return df

    def apply_exclusions_to_daily_gainers(self, df, tab_name):
        """
        对每日涨幅榜应用特殊的排除逻辑
        每日涨幅榜有29列格式：排名 + 14组(代币,日期)
        """
        original_count = len(df)
        
        if df.empty:
            return df
        
        # 获取需要排除的代币列表
        excluded_tokens = (self.stablecoins_to_exclude | 
                          self.major_coins_to_ignore | 
                          self.excluded_tokens)
        
        # 查找所有代币列 (代币, 代币_2, 代币_3, ... 代币_14)
        token_columns = []
        for col in df.columns:
            if col == '代币' or col.startswith('代币_'):
                token_columns.append(col)
        
        if not token_columns:
            print(f"  📊 {tab_name}: 未找到代币列，跳过排除逻辑")
            return df
        
        # 逐行检查并清理包含被排除代币的条目
        cleaned_rows = []
        for _, row in df.iterrows():
            cleaned_row = row.copy()
            
            # 检查每一组代币/日期对
            for i, token_col in enumerate(token_columns):
                date_col = '日期' if i == 0 else f'日期_{i+1}'
                
                # 如果代币在排除列表中，清空该组数据
                if (token_col in row and 
                    row[token_col] in excluded_tokens):
                    cleaned_row[token_col] = ''
                    if date_col in cleaned_row:
                        cleaned_row[date_col] = ''
            
            cleaned_rows.append(cleaned_row)
        
        cleaned_df = pd.DataFrame(cleaned_rows)
        
        # 统计清理的数量
        excluded_count = 0
        for token_col in token_columns:
            if token_col in df.columns:
                original_tokens = set(df[token_col].dropna())
                cleaned_tokens = set(cleaned_df[token_col].dropna())
                excluded_count += len(original_tokens - cleaned_tokens)
        
        if excluded_count > 0:
            print(f"  📊 {tab_name}: 清理了 {excluded_count} 个排除代币的条目")
        else:
            print(f"  📊 {tab_name}: 未发现需要排除的代币")
        
        return cleaned_df

    def create_output_folder(self):
        """创建输出文件夹"""
        timestamp = datetime.now().strftime('%Y%m%d')
        folder_name = f"币安代币分析结果_{timestamp}"

        # 创建主输出文件夹
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)
            print(f"📁 创建输出文件夹: {folder_name}")

        # 创建子文件夹
        subfolders = ['CSV文件', 'Excel文件', '分析报告', '日志文件']
        for subfolder in subfolders:
            subfolder_path = os.path.join(folder_name, subfolder)
            if not os.path.exists(subfolder_path):
                os.makedirs(subfolder_path)
                print(f"  📂 创建子文件夹: {subfolder}")

        return folder_name

    def get_output_path(self, filename, file_type='CSV文件'):
        """获取输出文件的完整路径"""
        return os.path.join(self.output_folder, file_type, filename)
    
    def display_version_info(self):
        """显示版本和功能信息"""
        print("🎯 币安代币筛选器 v2.0 - 生产版本")
        print("=" * 80)
        print(f"📅 发布日期: {self.release_date}")
        print(f"🔐 认证方式: OAuth (个人谷歌账户)")
        print(f"📊 数据来源: 币安API")
        print("=" * 80)
        print("✨ 核心功能:")
        print("  ✅ OAuth认证 (解决403存储配额错误)")
        print("  ✅ 用户可配置现货/期货代币数量")
        print("  ✅ 14日历史数据分析")
        print("  ✅ 实时资金费率")
        print("  ✅ 智能异常检测与颜色标注")
        print("  ✅ 每日独立谷歌表格文件")
        print("🆕 v2.0 核心功能:")
        print("  ✅ 修复OI显示为M/B格式")
        print("  ✅ 增强历史涨幅榜（29列格式）")
        print("  ✅ 准确的交易量数据计算")
        print("  ✅ 市值/期货交易量比值分析")
        print("  ✅ 全面的7日/14日涨跌数据")
        print("  ✅ 完整数据集保留（160+代币）")
        print("  ✅ 完整实现md.txt需求")
        print("=" * 80)
    
    def get_user_preferences(self):
        """获取用户偏好设置"""
        print("⚙️ 配置设置")
        print("-" * 50)

        # 如果是自动模式，使用默认值
        if self.auto_mode:
            print(f"🤖 自动模式：使用默认配置")
            print(f"  - 现货代币数量: {self.spot_count}")
            print(f"  - 期货代币数量: {self.futures_count}")
            return

        # 获取现货代币数量
        while True:
            try:
                spot_input = input(f"📈 请输入现货代币数量 (默认 {self.spot_count}): ").strip()
                if not spot_input:
                    break
                spot_count = int(spot_input)
                if spot_count <= 0:
                    print("❌ 数量必须大于0")
                    continue
                self.spot_count = spot_count
                break
            except ValueError:
                print("❌ 请输入有效数字")

        # 获取期货代币数量
        while True:
            try:
                futures_input = input(f"📊 请输入期货代币数量 (默认 {self.futures_count}): ").strip()
                if not futures_input:
                    break
                futures_count = int(futures_input)
                if futures_count <= 0:
                    print("❌ 数量必须大于0")
                    continue
                self.futures_count = futures_count
                break
            except ValueError:
                print("❌ 请输入有效数字")

        print(f"\n✅ 配置确认:")
        print(f"   现货代币: {self.spot_count}")
        print(f"   期货代币: {self.futures_count}")
        print()
    
    def generate_daily_sheet_name(self):
        """Generate daily sheet name"""
        today = datetime.now().strftime('%Y-%m-%d')
        return f"{self.base_sheet_name}_{today}"
    
    def authenticate_oauth_with_retry(self, max_retries=3):
        """OAuth认证 - 带重试机制和自动重新授权"""
        print("🔐 OAuth认证中...")

        if not os.path.exists(self.oauth_creds_file):
            print("❌ 未找到oauth_credentials.json文件")
            print("请先运行: python oauth_setup_v1.0.py")
            return False

        # 加载现有令牌
        if os.path.exists(self.token_file):
            try:
                self.creds = Credentials.from_authorized_user_file(self.token_file, self.SCOPES)
            except Exception as e:
                print(f"⚠️ 令牌文件损坏: {e}")
                print("将尝试重新授权...")
                return self.reauthorize()
        else:
            print("❌ 未找到token.json文件")
            print("将尝试自动授权...")
            return self.reauthorize()

        # 检查令牌状态
        if self.creds and self.creds.valid:
            print("✅ 现有令牌有效")
        elif self.creds and self.creds.expired and self.creds.refresh_token:
            print("🔄 令牌已过期，尝试刷新...")

            # 带重试的令牌刷新
            refresh_successful = False
            for attempt in range(max_retries):
                try:
                    print(f"🔄 刷新尝试 {attempt + 1}/{max_retries}...")
                    self.creds.refresh(Request())

                    # 保存刷新后的令牌
                    with open(self.token_file, 'w') as token:
                        token.write(self.creds.to_json())

                    print("✅ 令牌刷新成功")
                    refresh_successful = True
                    break

                except Exception as e:
                    error_msg = str(e).lower()
                    print(f"❌ 刷新尝试 {attempt + 1} 失败: {e}")
                    
                    # 检查是否是刷新令牌失效
                    if 'invalid_grant' in error_msg or 'token has been expired or revoked' in error_msg:
                        print("⚠️ 刷新令牌已失效，需要重新授权")
                        return self.reauthorize()

                    if attempt < max_retries - 1:
                        wait_time = (attempt + 1) * 2  # 递增等待时间
                        print(f"⏳ 等待 {wait_time} 秒后重试...")
                        time.sleep(wait_time)
            
            if not refresh_successful:
                print("❌ 所有刷新尝试都失败了")
                print("🔄 尝试重新授权...")
                return self.reauthorize()
                        
        else:
            print("❌ 令牌无效或缺失刷新令牌")
            print("🔄 尝试重新授权...")
            return self.reauthorize()

        # 初始化谷歌表格客户端
        try:
            self.gc = gspread.authorize(self.creds)
            print("✅ OAuth认证成功")
            return True
        except Exception as e:
            print(f"❌ 谷歌表格初始化失败: {e}")
            print("🔄 尝试重新授权...")
            return self.reauthorize()

    def reauthorize(self):
        """重新授权流程"""
        print("\n🔑 开始重新授权流程...")
        
        try:
            from google_auth_oauthlib.flow import InstalledAppFlow
            
            # 创建OAuth流程
            flow = InstalledAppFlow.from_client_secrets_file(
                self.oauth_creds_file, 
                self.SCOPES
            )
            
            print("📌 请在浏览器中完成授权...")
            print("提示: 请确保授予所有请求的权限")
            
            # 运行本地服务器接收授权
            self.creds = flow.run_local_server(
                port=8080,
                prompt='consent',  # 强制显示同意屏幕
                access_type='offline'  # 确保获取刷新令牌
            )
            
            # 保存新令牌
            with open(self.token_file, 'w') as token:
                token.write(self.creds.to_json())
            
            print("✅ 重新授权成功!")
            
            # 初始化谷歌表格客户端
            self.gc = gspread.authorize(self.creds)
            return True
            
        except Exception as e:
            print(f"❌ 重新授权失败: {e}")
            print("\n💡 请手动运行以下命令重新授权:")
            print("   python oauth_setup_enhanced.py")
            print("   或")
            print("   python oauth_setup_v1.0.py")
            return False
    
    def authenticate_oauth(self):
        """OAuth认证入口函数"""
        return self.authenticate_oauth_with_retry()
    
    def get_funding_rates(self):
        """获取期货资金费率"""
        print("💰 获取期货资金费率中...")
        funding_rates = {}

        try:
            response = requests.get("https://fapi.binance.com/fapi/v1/premiumIndex", timeout=10)
            if response.status_code == 200:
                data = response.json()

                for item in data:
                    symbol = item.get('symbol', '')
                    funding_rate = float(item.get('lastFundingRate', 0)) * 100
                    funding_rates[symbol] = funding_rate

                print(f"✅ 获取到 {len(funding_rates)} 个合约的资金费率")
            else:
                print(f"❌ 资金费率获取失败: {response.status_code}")

        except Exception as e:
            print(f"❌ 资金费率获取错误: {e}")

        return funding_rates

    def get_open_interest_data(self):
        """获取期货持仓量市值数据"""
        print("📊 获取持仓量市值数据中...")
        oi_data = {}
        futures_symbols = set()  # Track which symbols have futures

        try:
            # 直接使用24小时行情端点获取期货交易对列表
            response = requests.get("https://fapi.binance.com/fapi/v1/ticker/24hr", timeout=10)
            if response.status_code == 200:
                data = response.json()

                # 只获取USDT交易对的持仓量，避免过多请求
                usdt_symbols = [item['symbol'] for item in data if item['symbol'].endswith('USDT')]
                futures_symbols = set(usdt_symbols)  # Store all futures symbols
                
                # v1.1增强: 收集完整的期货市场数据用于缺失代币补充
                futures_market_data = {}
                for item in data:
                    if item['symbol'].endswith('USDT'):
                        futures_market_data[item['symbol']] = {
                            'symbol': item['symbol'],
                            'price': float(item.get('lastPrice', 0)),
                            'volume': float(item.get('volume', 0)),
                            'quoteVolume': float(item.get('quoteVolume', 0)),  # USD交易额
                            'priceChangePercent': float(item.get('priceChangePercent', 0)),
                            'count': int(item.get('count', 0)),
                            'highPrice': float(item.get('highPrice', 0)),
                            'lowPrice': float(item.get('lowPrice', 0)),
                            'openPrice': float(item.get('openPrice', 0))
                        }

                # 批量获取持仓量 - v1.1优化：优先处理主要代币
                processed = 0
                max_requests = 50  # v1.1: 平衡覆盖范围和API限制

                # 优先处理主要代币
                priority_symbols = ['BTCUSDT', 'ETHUSDT', 'SOLUSDT', 'XRPUSDT', 'BNBUSDT']
                priority_list = [s for s in priority_symbols if s in usdt_symbols]
                remaining_symbols = [s for s in usdt_symbols if s not in priority_symbols]
                
                # 合并优先列表和剩余列表
                ordered_symbols = priority_list + remaining_symbols[:max_requests-len(priority_list)]
                
                print(f"🔍 开始获取 {len(ordered_symbols)} 个合约的持仓量数据 (优先处理主要代币)...")

                for symbol in ordered_symbols:
                    try:
                        # 同时获取持仓量和价格
                        oi_response = requests.get(f"https://fapi.binance.com/fapi/v1/openInterest?symbol={symbol}", timeout=5)
                        price_response = requests.get(f"https://fapi.binance.com/fapi/v1/ticker/price?symbol={symbol}", timeout=5)

                        if oi_response.status_code == 200 and price_response.status_code == 200:
                            oi_info = oi_response.json()
                            price_info = price_response.json()

                            open_interest_contracts = float(oi_info.get('openInterest', 0))
                            price = float(price_info.get('price', 0))

                            # 计算OI市值
                            oi_market_value = open_interest_contracts * price
                            oi_data[symbol] = oi_market_value
                            processed += 1

                            # 显示重要代币的处理结果
                            if symbol in priority_symbols:
                                print(f"  ✅ {symbol}: OI市值 {oi_market_value/1000000:.1f}M")
                            elif processed % 10 == 0:
                                print(f"  📊 已处理 {processed} 个合约...")
                        else:
                            # 记录失败的请求，特别关注主要代币
                            if symbol in priority_symbols:
                                print(f"  ❌ 主要代币 {symbol} 获取失败!")
                                if oi_response.status_code != 200:
                                    print(f"      持仓量请求失败: {oi_response.status_code}")
                                if price_response.status_code != 200:
                                    print(f"      价格请求失败: {price_response.status_code}")
                            else:
                                if oi_response.status_code != 200:
                                    print(f"  ⚠️ {symbol} 持仓量请求失败: {oi_response.status_code}")
                                if price_response.status_code != 200:
                                    print(f"  ⚠️ {symbol} 价格请求失败: {price_response.status_code}")
                    except Exception as e:
                        print(f"  ⚠️ {symbol} 处理失败: {e}")
                        continue

                    # v1.1: 增强的速率限制 - 每3个请求休息更长时间
                    if processed % 3 == 0:
                        time.sleep(0.5)  # 增加延迟以避免被封禁

                print(f"✅ 获取到 {len(oi_data)} 个合约的持仓量市值")
                # Debug: 显示前5个OI数据
                sample_oi = list(oi_data.items())[:5]
                for symbol, oi_value in sample_oi:
                    print(f"  样例: {symbol} -> OI市值: {oi_value/1000000:.1f}M")
                
                # Debug: 检查主要代币是否在列表中
                major_futures = ['BTCUSDT', 'ETHUSDT', 'SOLUSDT', 'XRPUSDT', 'BNBUSDT']
                print(f"🔍 主要期货合约检查:")
                for symbol in major_futures:
                    if symbol in oi_data:
                        print(f"  ✅ {symbol}: {oi_data[symbol]/1000000:.1f}M")
                    elif symbol in usdt_symbols:
                        print(f"  ⚠️ {symbol}: 在期货列表中但OI获取失败")
                    else:
                        print(f"  ❌ {symbol}: 不在期货列表中")
            else:
                print(f"❌ 期货行情获取失败: {response.status_code}")

        except Exception as e:
            print(f"❌ 持仓量获取错误: {e}")
            futures_market_data = {}  # 确保在异常情况下也有默认值

        return oi_data, futures_symbols, futures_market_data

    def generate_final_report(self):
        """生成基础市场数据报告 - 分别获取现货和期货数据"""
        print("📊 获取币安市场数据中（分别获取现货和期货）...")

        try:
            # 1. 获取现货数据
            print(f"📈 获取前{self.spot_count}个现货交易对...")
            spot_url = "https://api.binance.com/api/v3/ticker/24hr"
            spot_response = requests.get(spot_url, timeout=10)

            if spot_response.status_code == 418:
                # Handle rate limit ban
                error_data = spot_response.json()
                print(f"⚠️ API访问被限制: {error_data.get('msg', 'Rate limit exceeded')}")
                print("💡 建议:")
                print("   1. 等待几分钟后重试")
                print("   2. 减少请求频率")
                print("   3. 使用WebSocket获取实时数据")
                return None
            elif spot_response.status_code != 200:
                print(f"❌ 获取现货数据失败 (状态码: {spot_response.status_code})")
                return None

            spot_data = spot_response.json()
            
            # 过滤USDT交易对并按交易量排序
            spot_usdt_pairs = [item for item in spot_data if item['symbol'].endswith('USDT')]
            spot_usdt_pairs.sort(key=lambda x: float(x['quoteVolume']), reverse=True)
            
            # 获取前N个现货交易对
            top_spot_pairs = spot_usdt_pairs[:min(self.spot_count, len(spot_usdt_pairs))]
            print(f"✅ 获取到前{len(top_spot_pairs)}个现货交易对")
            
            # 2. 获取期货数据
            print(f"📊 获取前{self.futures_count}个期货交易对...")
            futures_url = "https://fapi.binance.com/fapi/v1/ticker/24hr"
            futures_response = requests.get(futures_url, timeout=10)
            
            if futures_response.status_code != 200:
                print(f"❌ 获取期货数据失败 (状态码: {futures_response.status_code})")
                futures_data = []
            else:
                futures_data = futures_response.json()
            
            # 过滤USDT交易对并按交易量排序
            futures_usdt_pairs = [item for item in futures_data if item['symbol'].endswith('USDT')]
            futures_usdt_pairs.sort(key=lambda x: float(x['quoteVolume']), reverse=True)
            
            # 获取前N个期货交易对
            top_futures_pairs = futures_usdt_pairs[:min(self.futures_count, len(futures_usdt_pairs))]
            print(f"✅ 获取到前{len(top_futures_pairs)}个期货交易对")
            
            # 3. 构建数据字典用于合并
            spot_dict = {}
            futures_dict = {}
            
            # 处理现货数据
            for i, pair in enumerate(top_spot_pairs, 1):
                base_asset = pair['symbol'].replace('USDT', '')
                spot_dict[base_asset] = {
                    'spot_rank': i,
                    'base_asset': base_asset,
                    'symbol': pair['symbol'],
                    'price': float(pair['lastPrice']),
                    '1d_return': float(pair['priceChangePercent']),
                    'spot_volume_24h': float(pair['quoteVolume']),
                    'market_cap': float(pair['quoteVolume']) * 100,  # 简化的市值估算
                    'has_spot': True
                }
            
            # 处理期货数据
            for i, pair in enumerate(top_futures_pairs, 1):
                base_asset = pair['symbol'].replace('USDT', '')
                # 处理特殊代币（如1000PEPE -> PEPE）
                if base_asset.startswith('1000'):
                    base_asset = base_asset[4:]
                    
                futures_dict[base_asset] = {
                    'futures_rank': i,
                    'base_asset': base_asset,
                    'futures_symbol': pair['symbol'],
                    'futures_price': float(pair['lastPrice']),
                    'futures_1d_return': float(pair['priceChangePercent']),
                    'futures_volume_24h': float(pair['quoteVolume']),
                    'has_futures': True
                }
            
            # 4. 合并数据 - 以现货数据为主，期货数据为补充
            all_assets = set(spot_dict.keys()) | set(futures_dict.keys())
            print(f"📊 合并数据: 现货{len(spot_dict)}个 + 期货{len(futures_dict)}个 = 去重后{len(all_assets)}个代币")
            
            df_data = []
            for base_asset in all_assets:
                row_data = {
                    'base_asset': base_asset,
                    'has_spot': base_asset in spot_dict,
                    'has_futures': base_asset in futures_dict
                }
                
                # 优先使用现货数据
                if base_asset in spot_dict:
                    spot_info = spot_dict[base_asset]
                    row_data.update({
                        'volume_rank': spot_info['spot_rank'],
                        'symbol': spot_info['symbol'],
                        'price': spot_info['price'],
                        '1d_return': spot_info['1d_return'],
                        'spot_volume_24h': spot_info['spot_volume_24h'],
                        'market_cap': spot_info['market_cap']
                    })
                else:
                    # 没有现货数据，使用期货数据
                    futures_info = futures_dict[base_asset]
                    row_data.update({
                        'volume_rank': 999,  # 期货独有的代币排名靠后
                        'symbol': f"{base_asset}USDT",  # 构建标准符号
                        'price': futures_info['futures_price'],
                        '1d_return': futures_info['futures_1d_return'],
                        'spot_volume_24h': 0,  # 没有现货交易量
                        'market_cap': 0  # 期货独有代币暂无市值数据
                    })
                
                # 添加期货数据（如果有）
                if base_asset in futures_dict:
                    futures_info = futures_dict[base_asset]
                    row_data.update({
                        'futures_symbol': futures_info['futures_symbol'],
                        'futures_volume_24h': futures_info['futures_volume_24h']
                    })
                else:
                    row_data.update({
                        'futures_symbol': '',
                        'futures_volume_24h': 0
                    })
                
                # 计算总交易量
                row_data['total_volume_24h'] = row_data['spot_volume_24h'] + row_data['futures_volume_24h']
                
                df_data.append(row_data)
            
            df = pd.DataFrame(df_data)
            
            # 按总交易量重新排序
            df = df.sort_values('total_volume_24h', ascending=False).reset_index(drop=True)
            df['volume_rank'] = range(1, len(df) + 1)
            
            print(f"✅ 生成合并数据集:")
            print(f"   - 总代币数: {len(df)}")
            print(f"   - 仅现货: {len(df[df['has_spot'] & ~df['has_futures']])}个")
            print(f"   - 仅期货: {len(df[~df['has_spot'] & df['has_futures']])}个")
            print(f"   - 现货+期货: {len(df[df['has_spot'] & df['has_futures']])}个")
            
            return df

        except Exception as e:
            print(f"❌ 获取市场数据失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    def get_enhanced_market_data(self):
        """获取增强市场数据 - 根据7.15需求更新"""
        print(f"📊 获取增强市场数据中 (现货:{self.spot_count}, 期货:{self.futures_count})...")

        # 获取基础市场数据
        raw_df = self.generate_final_report()
        if raw_df is None or len(raw_df) == 0:
            print("❌ 获取基础市场数据失败")
            return None

        print(f"✅ 获取到 {len(raw_df)} 个代币的基础数据")

        # 根据7.15需求过滤排除的代币
        print("🔍 过滤排除的代币中...")
        initial_count = len(raw_df)
        raw_df = raw_df[~raw_df['base_asset'].isin(self.excluded_tokens)].copy()
        filtered_count = len(raw_df)
        print(f"✅ 过滤掉 {initial_count - filtered_count} 个排除的代币")

        # 获取资金费率
        funding_rates = self.get_funding_rates()

        # 获取持仓量数据
        oi_data, futures_symbols, futures_market_data = self.get_open_interest_data()
        
        # v1.1增强: 存储期货市场数据用于期货专注完整性增强
        self.futures_symbols = futures_symbols
        self.futures_market_data = futures_market_data
        
        # 获取真实市值数据从CoinGecko
        print("🦎 获取真实市值数据从CoinGecko...")
        all_symbols = raw_df['base_asset'].unique().tolist()
        market_cap_data = self.coingecko_client.get_binance_symbol_market_data(all_symbols)
        
        # 更新市值数据
        market_cap_updated = 0
        for idx, row in raw_df.iterrows():
            base_asset = row['base_asset']
            if base_asset in market_cap_data:
                coingecko_data = market_cap_data[base_asset]
                # 只有当CoinGecko有真实市值数据时才更新
                if coingecko_data.get('market_cap', 0) > 0:
                    raw_df.loc[idx, 'market_cap'] = coingecko_data['market_cap']
                    raw_df.loc[idx, 'circulating_supply'] = coingecko_data.get('circulating_supply', 0)
                    raw_df.loc[idx, 'market_cap_rank'] = coingecko_data.get('market_cap_rank', 0)
                    market_cap_updated += 1
                    
                    # 如果CoinGecko有7日/14日涨跌数据，也更新
                    if 'price_change_7d' in coingecko_data and coingecko_data['price_change_7d'] is not None:
                        raw_df.loc[idx, '7d_return'] = coingecko_data['price_change_7d']
                    if 'price_change_14d' in coingecko_data and coingecko_data['price_change_14d'] is not None:
                        raw_df.loc[idx, '14d_return'] = coingecko_data['price_change_14d']
        
        print(f"✅ 更新了{market_cap_updated}个代币的真实市值数据")

        # Initialize columns if they don't exist
        if 'funding_rate' not in raw_df.columns:
            raw_df['funding_rate'] = 0.0
        if 'open_interest' not in raw_df.columns:
            raw_df['open_interest'] = 0.0

        # Update has_futures based on actual futures data
        for idx, row in raw_df.iterrows():
            base_asset = row.get('base_asset', '')
            if base_asset in ['PEPE', 'BONK', 'FLOKI', 'SHIB', 'CAT']:
                symbol = f"1000{base_asset}USDT"
            else:
                symbol = f"{base_asset}USDT"
            
            raw_df.loc[idx, 'has_futures'] = symbol in futures_symbols

        # Add funding rates and OI to DataFrame
        print(f"📊 应用OI数据到DataFrame (共{len(oi_data)}个合约的OI数据)")
        print(f"📊 可用的OI符号样例: {list(oi_data.keys())[:5]}")
        oi_applied_count = 0
        missing_symbols = []
        applied_symbols = []
        for idx, row in raw_df.iterrows():
            if raw_df.loc[idx, 'has_futures']:
                base_asset = row.get('base_asset', '')
                if base_asset in ['PEPE', 'BONK', 'FLOKI', 'SHIB', 'CAT']:
                    symbol = f"1000{base_asset}USDT"
                else:
                    symbol = f"{base_asset}USDT"

                funding_rate = funding_rates.get(symbol, 0)
                open_interest = oi_data.get(symbol, 0)
                
                if open_interest > 0:
                    oi_applied_count += 1
                    applied_symbols.append(f"{base_asset}({symbol})")
                elif base_asset in ['BTC', 'ETH', 'SOL', 'XRP', 'BNB']:
                    missing_symbols.append(f"{base_asset}({symbol})")

                raw_df.loc[idx, 'funding_rate'] = funding_rate
                raw_df.loc[idx, 'open_interest'] = open_interest
                
                # 如果批量获取的OI为0，尝试单独获取
                if open_interest == 0:
                    try:
                        single_oi_response = requests.get(f"https://fapi.binance.com/fapi/v1/openInterest?symbol={symbol}", timeout=3)
                        single_price_response = requests.get(f"https://fapi.binance.com/fapi/v1/ticker/price?symbol={symbol}", timeout=3)
                        
                        if single_oi_response.status_code == 200 and single_price_response.status_code == 200:
                            single_oi_info = single_oi_response.json()
                            single_price_info = single_price_response.json()
                            
                            single_oi_contracts = float(single_oi_info.get('openInterest', 0))
                            single_price = float(single_price_info.get('price', 0))
                            single_oi_value = single_oi_contracts * single_price
                            
                            if single_oi_value > 0:
                                raw_df.loc[idx, 'open_interest'] = single_oi_value
                                print(f"  🔄 单独获取 {base_asset} OI成功: {single_oi_value/1000000:.1f}M")
                    except Exception as single_e:
                        pass  # 静默失败，保持原值
        
        print(f"✅ 成功应用OI数据到 {oi_applied_count} 个代币")
        if applied_symbols:
            print(f"✅ 有OI数据的代币: {', '.join(applied_symbols[:5])}")
        if missing_symbols:
            print(f"⚠️ 主要代币缺失OI数据: {', '.join(missing_symbols[:5])}")

        # Add 14-day historical data (from real API data)
        print("📊 计算14日历史数据中...")
        raw_df = self.add_real_historical_data(raw_df)

        # Add 7-day average OI (from real API data)
        print("📊 计算7日平均持仓量市值中...")
        raw_df = self.add_real_oi_averages(raw_df)

        # Get actual spot and futures volumes separately
        print("📊 获取实际现货和期货交易量数据...")
        raw_df = self.get_separate_spot_futures_volumes(raw_df)

        return raw_df

    def get_separate_spot_futures_volumes(self, df):
        """验证和调整现货期货交易量数据 - 数据已在generate_final_report中获取"""
        print("📊 验证现货和期货交易量数据...")
        
        try:
            # 验证必需的列是否存在
            required_columns = ['spot_volume_24h', 'futures_volume_24h', 'total_volume_24h']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                print(f"⚠️ 缺少列: {missing_columns}，正在添加...")
                for col in missing_columns:
                    df[col] = 0
            
            # 验证数据一致性
            inconsistent_count = 0
            for idx, row in df.iterrows():
                expected_total = row['spot_volume_24h'] + row['futures_volume_24h']
                actual_total = row['total_volume_24h']
                
                # 如果差异超过1%，认为不一致
                if abs(expected_total - actual_total) > actual_total * 0.01:
                    inconsistent_count += 1
                    # 修正总交易量
                    df.loc[idx, 'total_volume_24h'] = expected_total
            
            if inconsistent_count > 0:
                print(f"⚠️ 修正了 {inconsistent_count} 条数据的总交易量")
            
            # 显示数据统计
            spot_only = len(df[(df['spot_volume_24h'] > 0) & (df['futures_volume_24h'] == 0)])
            futures_only = len(df[(df['spot_volume_24h'] == 0) & (df['futures_volume_24h'] > 0)])
            both = len(df[(df['spot_volume_24h'] > 0) & (df['futures_volume_24h'] > 0)])
            
            print(f"✅ 交易量数据验证完成:")
            print(f"   - 仅现货交易: {spot_only}个")
            print(f"   - 仅期货交易: {futures_only}个")
            print(f"   - 现货+期货交易: {both}个")
            
        except Exception as e:
            print(f"❌ 验证交易量数据失败: {e}")
            import traceback
            traceback.print_exc()
        
        return df

    def smart_format_oi(self, x):
        """智能格式化持仓量市值 - v1.1 修复显示为M/B格式"""
        if pd.isna(x) or x <= 0:
            return "N/A"

        # 持仓量市值使用M/B单位显示（不带美元符号）
        if x >= 1000000000:  # 10亿以上
            return f"{x/1000000000:.1f}B"
        elif x >= 1000000:   # 100万以上
            return f"{x/1000000:.1f}M"
        elif x >= 1000:      # 1000以上
            return f"{x/1000:.1f}K"
        else:
            return f"{x:.0f}"

    def get_historical_klines(self, symbol, interval='1d', limit=14, is_futures=False):
        """获取历史K线数据"""
        try:
            if is_futures:
                url = "https://fapi.binance.com/fapi/v1/klines"
            else:
                url = "https://api.binance.com/api/v3/klines"

            params = {
                'symbol': symbol,
                'interval': interval,
                'limit': limit
            }

            response = requests.get(url, params=params, timeout=10)
            if response.status_code == 200:
                data = response.json()
                klines = []
                for item in data:
                    klines.append({
                        'timestamp': item[0],
                        'open': float(item[1]),
                        'high': float(item[2]),
                        'low': float(item[3]),
                        'close': float(item[4]),
                        'volume': float(item[5]),
                        'quote_volume': float(item[7])
                    })
                return klines
            return []
        except Exception as e:
            print(f"  ⚠️ 获取{symbol}历史数据失败: {e}")
            return []

    def get_oi_history(self, symbol, hours=168):
        """获取持仓量历史数据 (默认7天=168小时)"""
        try:
            url = "https://fapi.binance.com/futures/data/openInterestHist"
            params = {
                'symbol': symbol,
                'period': '1h',
                'limit': hours
            }

            response = requests.get(url, params=params, timeout=10)
            if response.status_code == 200:
                data = response.json()
                return [float(item['sumOpenInterest']) for item in data]
            return []
        except Exception as e:
            print(f"  ⚠️ 获取{symbol}持仓量历史失败: {e}")
            return []

    def add_real_historical_data(self, df):
        """添加真实的14日历史数据 - 修复：无现货时使用期货数据计算涨跌"""
        for idx, row in df.iterrows():
            base_asset = row.get('base_asset', '')
            if not base_asset:
                continue

            # 添加CAT的特殊映射
            if base_asset == 'CAT':
                futures_symbol_override = '1000CATUSDT'
            else:
                futures_symbol_override = None

            # 标记是否已计算涨跌数据
            has_price_data = False

            # 现货历史数据
            if row.get('has_spot', False):
                spot_symbol = f"{base_asset}USDT"
                spot_klines = self.get_historical_klines(spot_symbol, '1d', 15, is_futures=False)

                if spot_klines and len(spot_klines) >= 14:
                    # 计算14日平均现货交易量
                    spot_volumes = [k['quote_volume'] for k in spot_klines]
                    avg_spot_volume_14d = sum(spot_volumes) / len(spot_volumes)
                    df.loc[idx, 'avg_spot_volume_14d'] = avg_spot_volume_14d

                    # 计算14日收益率
                    if len(spot_klines) >= 15:
                        price_14d_ago = spot_klines[0]['close']
                        current_price = spot_klines[-1]['close']
                        return_14d = ((current_price / price_14d_ago) - 1) * 100
                        df.loc[idx, '14d_return'] = return_14d
                        has_price_data = True
                    
                    # 计算7日收益率
                    if len(spot_klines) >= 8:
                        price_7d_ago = spot_klines[-8]['close']  # 7天前的收盘价
                        current_price = spot_klines[-1]['close']
                        return_7d = ((current_price / price_7d_ago) - 1) * 100
                        df.loc[idx, '7d_return'] = return_7d
                    else:
                        df.loc[idx, '7d_return'] = 0
                else:
                    df.loc[idx, 'avg_spot_volume_14d'] = 0
                    if not has_price_data:
                        df.loc[idx, '14d_return'] = 0
                        df.loc[idx, '7d_return'] = 0

            # 期货历史数据
            if row.get('has_futures', False):
                # 处理特殊合约映射
                if futures_symbol_override:
                    futures_symbol = futures_symbol_override
                elif base_asset in ['PEPE', 'SHIB', 'FLOKI', 'BONK', 'CAT']:
                    futures_symbol = f"1000{base_asset}USDT"
                else:
                    futures_symbol = f"{base_asset}USDT"

                futures_klines = self.get_historical_klines(futures_symbol, '1d', 15, is_futures=True)

                if futures_klines and len(futures_klines) >= 14:
                    # 计算14日平均期货交易量
                    futures_volumes = [k['quote_volume'] for k in futures_klines]
                    avg_futures_volume_14d = sum(futures_volumes) / len(futures_volumes)
                    df.loc[idx, 'avg_futures_volume_14d'] = avg_futures_volume_14d
                    
                    # 如果没有现货数据，使用期货数据计算涨跌
                    if not has_price_data:
                        # 计算14日收益率
                        if len(futures_klines) >= 15:
                            price_14d_ago = futures_klines[0]['close']
                            current_price = futures_klines[-1]['close']
                            return_14d = ((current_price / price_14d_ago) - 1) * 100
                            df.loc[idx, '14d_return'] = return_14d
                        
                        # 计算7日收益率
                        if len(futures_klines) >= 8:
                            price_7d_ago = futures_klines[-8]['close']  # 7天前的收盘价
                            current_price = futures_klines[-1]['close']
                            return_7d = ((current_price / price_7d_ago) - 1) * 100
                            df.loc[idx, '7d_return'] = return_7d
                        else:
                            df.loc[idx, '7d_return'] = 0
                else:
                    df.loc[idx, 'avg_futures_volume_14d'] = 0
                    if not has_price_data:
                        df.loc[idx, '14d_return'] = 0
                        df.loc[idx, '7d_return'] = 0

        return df

    def add_real_oi_averages(self, df):
        """添加真实的7日平均持仓量市值"""
        for idx, row in df.iterrows():
            base_asset = row.get('base_asset', '')
            if not base_asset or not row.get('has_futures', False):
                df.loc[idx, 'avg_oi_7d'] = 0
                continue

            # 处理特殊合约映射
            if base_asset in ['PEPE', 'SHIB', 'FLOKI', 'BONK', 'CAT']:
                futures_symbol = f"1000{base_asset}USDT"
            else:
                futures_symbol = f"{base_asset}USDT"

            # 获取7日持仓量历史和价格历史
            oi_market_value_7d = self.get_oi_market_value_history(futures_symbol, 168)  # 7天

            if oi_market_value_7d and len(oi_market_value_7d) > 0:
                avg_oi_market_value_7d = sum(oi_market_value_7d) / len(oi_market_value_7d)
                df.loc[idx, 'avg_oi_7d'] = avg_oi_market_value_7d
            else:
                # 如果没有历史数据，使用当前持仓量市值
                df.loc[idx, 'avg_oi_7d'] = row.get('open_interest', 0)

        return df

    def get_oi_market_value_history(self, symbol, hours=168):
        """获取持仓量市值历史数据 (默认7天=168小时)"""
        try:
            # 获取持仓量历史
            oi_url = "https://fapi.binance.com/futures/data/openInterestHist"
            oi_params = {
                'symbol': symbol,
                'period': '1h',
                'limit': hours
            }

            oi_response = requests.get(oi_url, params=oi_params, timeout=10)
            if oi_response.status_code != 200:
                return []

            oi_data = oi_response.json()

            # 获取价格历史 (K线数据)
            price_url = "https://fapi.binance.com/fapi/v1/klines"
            price_params = {
                'symbol': symbol,
                'interval': '1h',
                'limit': hours
            }

            price_response = requests.get(price_url, params=price_params, timeout=10)
            if price_response.status_code != 200:
                return []

            price_data = price_response.json()

            # 计算每小时的OI市值
            oi_market_values = []
            min_length = min(len(oi_data), len(price_data))

            for i in range(min_length):
                oi_contracts = float(oi_data[i]['sumOpenInterest'])
                price = float(price_data[i][4])  # 收盘价
                oi_market_value = oi_contracts * price
                oi_market_values.append(oi_market_value)

            return oi_market_values

        except Exception as e:
            print(f"  ⚠️ 获取{symbol}持仓量市值历史失败: {e}")
            return []

    def detect_refined_anomalies(self, df, dataset_key):
        """Detect refined anomalies per 7.15 requirements - single cell marking"""
        print(f"🔍 Detecting refined anomalies for {dataset_key}...")

        # Special handling for daily_gainers_history which has different structure
        if dataset_key == 'daily_gainers_history':
            # No anomaly detection for historical gainers data
            return {}

        # Filter out major coins for anomaly detection (except raw_data)
        if dataset_key != 'raw_data' and 'base_asset' in df.columns:
            analysis_df = df[~df['base_asset'].isin(self.major_coins_to_ignore)].copy()
        else:
            analysis_df = df.copy()

        anomaly_info = {
            'futures_volume_growth_top5': [],
            'spot_volume_growth_top5': [],
            '1d_return_top5': [],
            '14d_return_top5': [],
            'volume_vs_mcap_top5': [],
            'funding_rate_anomaly': [],
            'oi_vs_mcap': [],
            'oi_vs_avg_oi': []
        }

        # 1. Futures volume growth vs 14d average (top 5) - 修复计算逻辑
        if 'futures_volume_24h' in analysis_df.columns and 'avg_futures_volume_14d' in analysis_df.columns:
            # 只对有历史数据的代币计算比例
            valid_data = analysis_df[analysis_df['avg_futures_volume_14d'] > 0].copy()
            if len(valid_data) > 0:
                valid_data['futures_growth_ratio'] = valid_data['futures_volume_24h'] / valid_data['avg_futures_volume_14d']
                top5_futures = valid_data.nlargest(5, 'futures_growth_ratio')
                anomaly_info['futures_volume_growth_top5'] = top5_futures.index.tolist()

        # 2. Spot volume growth vs 14d average (top 5) - 修复计算逻辑
        if 'spot_volume_24h' in analysis_df.columns and 'avg_spot_volume_14d' in analysis_df.columns:
            # 只对有历史数据的代币计算比例
            valid_data = analysis_df[analysis_df['avg_spot_volume_14d'] > 0].copy()
            if len(valid_data) > 0:
                valid_data['spot_growth_ratio'] = valid_data['spot_volume_24h'] / valid_data['avg_spot_volume_14d']
                top5_spot = valid_data.nlargest(5, 'spot_growth_ratio')
                anomaly_info['spot_volume_growth_top5'] = top5_spot.index.tolist()

        # 3. 1-day return top 5 (both positive and negative)
        if '1d_return' in analysis_df.columns:
            # 确保数据是数值类型
            analysis_df['1d_return'] = pd.to_numeric(analysis_df['1d_return'], errors='coerce')
            valid_1d = analysis_df.dropna(subset=['1d_return'])
            if len(valid_1d) > 0:
                top5_gainers = valid_1d.nlargest(5, '1d_return')
                top5_losers = valid_1d.nsmallest(5, '1d_return')
                anomaly_info['1d_return_top5'] = top5_gainers.index.tolist() + top5_losers.index.tolist()

        # 4. 14-day return top 5 (both positive and negative)
        if '14d_return' in analysis_df.columns:
            # 确保数据是数值类型
            analysis_df['14d_return'] = pd.to_numeric(analysis_df['14d_return'], errors='coerce')
            valid_14d = analysis_df.dropna(subset=['14d_return'])
            if len(valid_14d) > 0:
                top5_14d_gainers = valid_14d.nlargest(5, '14d_return')
                top5_14d_losers = valid_14d.nsmallest(5, '14d_return')
                anomaly_info['14d_return_top5'] = top5_14d_gainers.index.tolist() + top5_14d_losers.index.tolist()

        # 5. Volume vs market cap (extremely high relative to market cap)
        if 'total_volume_24h' in analysis_df.columns and 'market_cap' in analysis_df.columns:
            # 确保数据是数值类型
            analysis_df['total_volume_24h'] = pd.to_numeric(analysis_df['total_volume_24h'], errors='coerce')
            analysis_df['market_cap'] = pd.to_numeric(analysis_df['market_cap'], errors='coerce')
            valid_volume_mcap = analysis_df.dropna(subset=['total_volume_24h', 'market_cap'])
            if len(valid_volume_mcap) > 0:
                valid_volume_mcap['volume_mcap_ratio'] = valid_volume_mcap['total_volume_24h'] / (valid_volume_mcap['market_cap'] + 1)
                top5_volume_mcap = valid_volume_mcap.nlargest(5, 'volume_mcap_ratio')
                anomaly_info['volume_vs_mcap_top5'] = top5_volume_mcap.index.tolist()

        # 6. Funding rate anomalies (>0.1% or <-0.1%)
        if 'funding_rate' in analysis_df.columns:
            # 确保数据是数值类型
            analysis_df['funding_rate'] = pd.to_numeric(analysis_df['funding_rate'], errors='coerce')
            valid_funding = analysis_df.dropna(subset=['funding_rate'])
            if len(valid_funding) > 0:
                funding_anomalies = valid_funding[abs(valid_funding['funding_rate']) > 0.1]
                anomaly_info['funding_rate_anomaly'] = funding_anomalies.index.tolist()

        # 7. OI vs market cap (for futures focus)
        if dataset_key == 'futures_focus' and 'open_interest' in analysis_df.columns and 'market_cap' in analysis_df.columns:
            # 确保数据是数值类型
            analysis_df['open_interest'] = pd.to_numeric(analysis_df['open_interest'], errors='coerce')
            analysis_df['market_cap'] = pd.to_numeric(analysis_df['market_cap'], errors='coerce')
            valid_oi_mcap = analysis_df.dropna(subset=['open_interest', 'market_cap'])
            if len(valid_oi_mcap) > 0:
                valid_oi_mcap['oi_mcap_ratio'] = valid_oi_mcap['open_interest'] / (valid_oi_mcap['market_cap'] + 1)
                top5_oi_mcap = valid_oi_mcap.nlargest(5, 'oi_mcap_ratio')
                anomaly_info['oi_vs_mcap'] = top5_oi_mcap.index.tolist()

        # 8. OI vs 7d average OI
        if dataset_key == 'futures_focus' and 'open_interest' in analysis_df.columns and 'avg_oi_7d' in analysis_df.columns:
            # 确保数据是数值类型
            analysis_df['open_interest'] = pd.to_numeric(analysis_df['open_interest'], errors='coerce')
            analysis_df['avg_oi_7d'] = pd.to_numeric(analysis_df['avg_oi_7d'], errors='coerce')
            valid_oi_growth = analysis_df.dropna(subset=['open_interest', 'avg_oi_7d'])
            if len(valid_oi_growth) > 0:
                valid_oi_growth['oi_growth_ratio'] = valid_oi_growth['open_interest'] / (valid_oi_growth['avg_oi_7d'] + 1)
                top5_oi_growth = valid_oi_growth.nlargest(5, 'oi_growth_ratio')
                anomaly_info['oi_vs_avg_oi'] = top5_oi_growth.index.tolist()

        total_anomalies = sum(len(indices) for indices in anomaly_info.values())
        print(f"✅ Detected {total_anomalies} refined anomalies for {dataset_key}")

        return anomaly_info

    def format_data_for_sheets(self, df):
        """Format data for Google Sheets display"""
        formatted_df = df.copy()

        # Format numeric columns
        numeric_formatters = {
            'price': lambda x: f"{x:.6f}" if pd.notna(x) and x < 1 else f"{x:.4f}" if pd.notna(x) else "N/A",
            'market_cap': lambda x: f"{x/1000000:.1f}M" if pd.notna(x) and x > 0 else "N/A",
            'total_volume_24h': lambda x: f"{x/1000000:.1f}M" if pd.notna(x) and x > 0 else "N/A",
            'spot_volume_24h': lambda x: f"{x/1000000:.1f}M" if pd.notna(x) and x > 0 else "N/A",
            'futures_volume_24h': lambda x: f"{x/1000000:.1f}M" if pd.notna(x) and x > 0 else "N/A",
            'avg_spot_volume_14d': lambda x: f"{x/1000000:.1f}M" if pd.notna(x) and x > 0 else "N/A",
            'avg_futures_volume_14d': lambda x: f"{x/1000000:.1f}M" if pd.notna(x) and x > 0 else "N/A",
            'open_interest': lambda x: self.smart_format_oi(x),
            'avg_oi_7d': lambda x: self.smart_format_oi(x),
            'mcap_futures_ratio': lambda x: f"{x:.2f}" if pd.notna(x) and x > 0 else "N/A",
            '1d_return': lambda x: f"{x:+.2f}%" if pd.notna(x) else "N/A",
            '7d_return': lambda x: f"{x:+.2f}%" if pd.notna(x) else "N/A",
            '14d_return': lambda x: f"{x:+.2f}%" if pd.notna(x) else "N/A",
            'funding_rate': lambda x: f"{x:+.4f}%" if pd.notna(x) else "N/A"
        }

        for col, formatter in numeric_formatters.items():
            if col in formatted_df.columns:
                formatted_df[col] = formatted_df[col].apply(formatter)

        # Format boolean columns
        bool_columns = ['has_spot', 'has_futures']
        for col in bool_columns:
            if col in formatted_df.columns:
                formatted_df[col] = formatted_df[col].apply(lambda x: "✅" if x else "❌")

        return formatted_df

    def create_enhanced_datasets(self, df):
        """Create enhanced datasets for analysis - Updated per 7.28 requirements"""
        print("🔍 Creating enhanced datasets with selective exclusions...")

        datasets = {}
        
        # Create tab name mapping for Chinese names
        tab_name_mapping = {
            'raw_data': '原始数据',
            'volume_ranking': '交易量排行',
            'market_cap_ranking': '市值排行',
            'gainers_losers': '涨跌排行',
            'futures_focus': '期货专注',
            'daily_gainers_history': '每日涨幅榜',
            'low_volume_high_mcap': '低量高市值',
            'high_volume_low_mcap': '高市值低期货量',
            'low_mcap_high_volume': '低市值高交易量'
        }

        # 1. Raw data - 核心标签页，排除稳定币但保留主要币种
        raw_data = df.copy()
        raw_data = self.apply_selective_exclusions(raw_data, tab_name_mapping['raw_data'])
        datasets['raw_data'] = raw_data

        # 2. Volume ranking - 核心标签页，排除稳定币但保留主要币种
        volume_df = df.copy().sort_values('total_volume_24h', ascending=False).reset_index(drop=True)
        volume_df['volume_rank'] = range(1, len(volume_df) + 1)
        volume_df = self.apply_selective_exclusions(volume_df, tab_name_mapping['volume_ranking'])
        datasets['volume_ranking'] = volume_df

        # 3. Market cap ranking - 分析标签页，排除稳定币和主要币种 (7.28需求4更新)
        market_cap_df = df[df['market_cap'] > 0].copy()
        market_cap_df = market_cap_df.sort_values('market_cap', ascending=False).reset_index(drop=True)
        market_cap_df['mcap_rank'] = range(1, len(market_cap_df) + 1)
        market_cap_df = self.apply_selective_exclusions(market_cap_df, tab_name_mapping['market_cap_ranking'])
        datasets['market_cap_ranking'] = market_cap_df

        # 4. Gainers and losers - 分析标签页，排除稳定币和主要币种
        gainers = df[df['1d_return'] > 0].nlargest(30, '1d_return').copy()
        losers = df[df['1d_return'] < 0].nsmallest(30, '1d_return').copy()
        gainers['gain_loss_type'] = '涨幅榜'
        losers['gain_loss_type'] = '跌幅榜'
        gainers_losers = pd.concat([gainers, losers]).reset_index(drop=True)
        gainers_losers['gain_loss_rank'] = range(1, len(gainers_losers) + 1)
        gainers_losers = self.apply_selective_exclusions(gainers_losers, tab_name_mapping['gainers_losers'])
        datasets['gainers_losers'] = gainers_losers

        # 5. Futures focus - 核心标签页，排除稳定币但保留主要币种
        if hasattr(self, 'futures_symbols') and hasattr(self, 'futures_market_data'):
            futures_df = self.create_comprehensive_futures_focus(
                df, self.futures_symbols, self.futures_market_data
            )
        else:
            # 备选方案: 使用原有逻辑
            print("⚠️  使用原有期货专注逻辑 (缺少增强数据)")
            futures_df = df[df['has_futures'] == True].copy()
            futures_df = futures_df.sort_values('futures_volume_24h', ascending=False).reset_index(drop=True)
            futures_df['futures_rank'] = range(1, len(futures_df) + 1)
        futures_df = self.apply_selective_exclusions(futures_df, tab_name_mapping['futures_focus'])
        datasets['futures_focus'] = futures_df

        # 6. Daily gainers history - 分析标签页，排除稳定币和主要币种
        daily_gainers = self.create_daily_gainers_history(df)
        daily_gainers = self.apply_selective_exclusions(daily_gainers, tab_name_mapping['daily_gainers_history'])
        datasets['daily_gainers_history'] = daily_gainers

        # 7. High market cap low futures volume - 修改逻辑：市值前100中期货交易额最少的35个
        # 首先获取市值前100的代币
        top_100_mcap = df[df['market_cap'] > 0].nlargest(100, 'market_cap')
        
        # 在这100个代币中，筛选有期货交易的代币
        top_100_with_futures = top_100_mcap[top_100_mcap['futures_volume_24h'] > 0].copy()
        
        # 按期货交易额升序排序，取最少的35个
        high_mcap_low_vol = top_100_with_futures.nsmallest(35, 'futures_volume_24h').copy()
        high_mcap_low_vol = high_mcap_low_vol.reset_index(drop=True)
        
        # 添加市值/期货交易量比值供参考
        high_mcap_low_vol['mcap_futures_ratio'] = high_mcap_low_vol['market_cap'] / (high_mcap_low_vol['futures_volume_24h'] / 1000000)
        
        high_mcap_low_vol = self.apply_selective_exclusions(high_mcap_low_vol, tab_name_mapping['high_volume_low_mcap'])
        datasets['high_volume_low_mcap'] = high_mcap_low_vol

        # 8. 低量高市值 - 市值前150中期货交易量最低的35个 (完整实现)
        print("🔍 创建低量高市值tab - 完整实现...")
        
        # Step 1: 从CoinGecko获取真实的市值前150代币
        top_150_symbols, top_150_market_data = self.coingecko_client.get_top_market_cap_symbols(150)
        print(f"🦎 从CoinGecko获取到市值前{len(top_150_symbols)}的代币")
        
        # Step 2: 获取这150个代币的期货交易量
        print("📊 获取市值前150代币的期货交易量...")
        
        # 先从已有数据中查找
        available_in_df = df[df['base_asset'].isin(top_150_symbols)].copy()
        available_symbols = set(available_in_df['base_asset'].unique())
        missing_symbols = list(set(top_150_symbols) - available_symbols)
        
        print(f"  - 已有数据: {len(available_symbols)}个")
        print(f"  - 需要补充: {len(missing_symbols)}个")
        
        # Step 3: 补充缺失代币的数据
        if missing_symbols:
            print(f"📊 补充{len(missing_symbols)}个缺失代币的数据...")
            
            # 使用数据补全器获取缺失代币的数据
            supplemented_df = self.data_supplementer.supplement_multiple_tokens(
                missing_symbols[:20],  # 限制一次最多补充20个，避免超时
                top_150_market_data
            )
            
            # 合并补充的数据
            if not supplemented_df.empty:
                # 确保列结构一致
                for col in df.columns:
                    if col not in supplemented_df.columns:
                        supplemented_df[col] = 0
                
                # 合并数据
                df = pd.concat([df, supplemented_df], ignore_index=True)
                print(f"✅ 成功补充{len(supplemented_df)}个代币的数据")
                
                # 更新available_in_df
                available_in_df = df[df['base_asset'].isin(top_150_symbols)].copy()
        
        # Step 4: 筛选有期货交易的代币
        top_mcap_with_futures = available_in_df[available_in_df['futures_volume_24h'] > 0].copy()
        
        # 确保使用真实市值数据
        for idx, row in top_mcap_with_futures.iterrows():
            symbol = row['base_asset']
            if symbol in top_150_market_data:
                real_mcap = top_150_market_data[symbol].get('market_cap', row['market_cap'])
                if real_mcap > 0:
                    top_mcap_with_futures.loc[idx, 'market_cap'] = real_mcap
                    top_mcap_with_futures.loc[idx, 'market_cap_rank'] = top_150_market_data[symbol].get('market_cap_rank', 0)
        
        # Step 5: 按期货交易量升序排序，取最少的35个
        print(f"📊 从{len(top_mcap_with_futures)}个有期货的高市值代币中选择...")
        
        # 排除主要币种（BTC, ETH, BNB, SOL, XRP）- 这些不应该出现在低量高市值中
        top_mcap_with_futures_filtered = top_mcap_with_futures[~top_mcap_with_futures['base_asset'].isin(self.major_coins_to_ignore)].copy()
        print(f"  - 排除主要币种后: {len(top_mcap_with_futures_filtered)}个")
        
        low_vol_high_mcap = top_mcap_with_futures_filtered.nsmallest(35, 'futures_volume_24h').copy()
        low_vol_high_mcap = low_vol_high_mcap.sort_values('futures_volume_24h', ascending=True).reset_index(drop=True)
        
        # 添加排名
        low_vol_high_mcap['low_vol_rank'] = range(1, len(low_vol_high_mcap) + 1)
        
        print(f"✅ 低量高市值tab: 选出期货交易量最低的{len(low_vol_high_mcap)}个代币")
        
        # 准备低量高市值tab所需的特定列
        low_vol_high_mcap_formatted = pd.DataFrame()
        low_vol_high_mcap_formatted['代币'] = low_vol_high_mcap['base_asset']
        low_vol_high_mcap_formatted['价格(USDT)'] = low_vol_high_mcap['price'].round(6)
        low_vol_high_mcap_formatted['市值(M USD)'] = (low_vol_high_mcap['market_cap'] / 1_000_000).round(2)
        low_vol_high_mcap_formatted['24h总量(M USDT)'] = (low_vol_high_mcap['total_volume_24h'] / 1_000_000).round(2)
        low_vol_high_mcap_formatted['24h现货交易量(M USDT)'] = (low_vol_high_mcap['spot_volume_24h'] / 1_000_000).round(2)
        # Handle 14-day averages properly - check column existence first
        if 'avg_spot_volume_14d' in low_vol_high_mcap.columns:
            low_vol_high_mcap_formatted['14日平均现货交易量(M USDT)'] = low_vol_high_mcap['avg_spot_volume_14d'].fillna(0).apply(lambda x: round(x / 1_000_000, 2))
        else:
            low_vol_high_mcap_formatted['14日平均现货交易量(M USDT)'] = 0
            
        low_vol_high_mcap_formatted['24h合约交易量(M USDT)'] = (low_vol_high_mcap['futures_volume_24h'] / 1_000_000).round(2)
        
        # Handle futures 14-day average properly
        if 'avg_futures_volume_14d' in low_vol_high_mcap.columns:
            low_vol_high_mcap_formatted['14日期货合约交易量(M USDT)'] = low_vol_high_mcap['avg_futures_volume_14d'].fillna(0).apply(lambda x: round(x / 1_000_000, 2))
        else:
            low_vol_high_mcap_formatted['14日期货合约交易量(M USDT)'] = 0
        low_vol_high_mcap_formatted['1日涨跌(%)'] = low_vol_high_mcap['1d_return'].round(2)
        low_vol_high_mcap_formatted['7日涨跌'] = low_vol_high_mcap['7d_return'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "N/A") if '7d_return' in low_vol_high_mcap.columns else "N/A"
        low_vol_high_mcap_formatted['14日涨跌'] = low_vol_high_mcap['14d_return'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "N/A") if '14d_return' in low_vol_high_mcap.columns else "N/A"
        low_vol_high_mcap_formatted['资金费率(%)'] = low_vol_high_mcap['funding_rate'].apply(lambda x: f"{x:.4f}" if pd.notna(x) else "N/A") if 'funding_rate' in low_vol_high_mcap.columns else "N/A"
        
        # Add base_asset column for exclusion logic
        low_vol_high_mcap_formatted['base_asset'] = low_vol_high_mcap['base_asset']
        low_vol_high_mcap_formatted = self.apply_selective_exclusions(low_vol_high_mcap_formatted, tab_name_mapping['low_volume_high_mcap'])
        
        # Remove base_asset column after exclusion (not needed in final output)
        if 'base_asset' in low_vol_high_mcap_formatted.columns:
            low_vol_high_mcap_formatted = low_vol_high_mcap_formatted.drop('base_asset', axis=1)
        datasets['low_volume_high_mcap'] = low_vol_high_mcap_formatted

        # 9. 低市值高交易量 - 市值/合约交易量 比值最低的top30
        # 筛选有期货交易量的代币
        df_with_futures_for_low = df[df['futures_volume_24h'] > 0].copy()
        # 计算市值/期货交易量比值
        df_with_futures_for_low['mcap_futures_ratio'] = df_with_futures_for_low['market_cap'] / (df_with_futures_for_low['futures_volume_24h'] / 1000000)
        
        # 低市值高交易量: 市值/合约交易量 比值最低的top30
        low_mcap_high_volume = df_with_futures_for_low.nsmallest(30, 'mcap_futures_ratio').copy()
        low_mcap_high_volume = low_mcap_high_volume.reset_index(drop=True)
        low_mcap_high_volume = self.apply_selective_exclusions(low_mcap_high_volume, '低市值高交易量')
        datasets['low_mcap_high_volume'] = low_mcap_high_volume

        print(f"✅ Created {len(datasets)} enhanced datasets with selective exclusions")
        return datasets

    def create_daily_gainers_history(self, df):
        """创建每日涨幅榜历史数据 - 按照用户要求的格式"""
        print("📈 创建每日涨幅榜历史数据...")

        # 获取当前日期的涨幅榜前20名
        current_gainers = df[df['1d_return'] > 0].nlargest(20, '1d_return').copy()
        current_date = datetime.now().strftime('%m-%d')  # 7.28需求5: 只显示月日

        # 尝试读取历史数据
        history_data = self.load_daily_gainers_history()

        # 创建当前日期的数据
        current_data = []
        for i, (_, row) in enumerate(current_gainers.iterrows(), 1):
            current_data.append({
                '排名': i,
                '代币': row['base_asset'],
                '日期': current_date,
                '代币_2': row['base_asset'],  # 第二列代币
                '日期_2': current_date,      # 第二列日期
                '代币_3': row['base_asset'],  # 第三列代币
                '日期_3': current_date       # 第三列日期
            })

        # 合并历史数据和当前数据
        combined_data = self.merge_historical_gainers_data(history_data, current_data, current_date)

        # 保存更新后的历史数据
        self.save_daily_gainers_history(combined_data)

        # 转换为DataFrame格式用于上传
        formatted_df = self.format_gainers_history_for_display(combined_data)

        print(f"✅ 每日涨幅榜历史数据已更新，包含 {len(formatted_df)} 行数据")
        return formatted_df

    def load_daily_gainers_history(self):
        """加载历史涨幅榜数据"""
        history_file = 'daily_gainers_history.json'

        try:
            if os.path.exists(history_file):
                with open(history_file, 'r', encoding='utf-8') as f:
                    history_data = json.load(f)
                print(f"📚 加载历史数据: {len(history_data)} 个日期的记录")
                return history_data
            else:
                print("📚 未找到历史数据文件，创建新的历史记录")
                return {}
        except Exception as e:
            print(f"⚠️ 加载历史数据失败: {e}")
            return {}

    def merge_historical_gainers_data(self, history_data, current_data, current_date):
        """合并历史数据和当前数据，按照用户要求的14天格式"""
        # 更新历史数据
        history_data[current_date] = current_data

        # 获取最近的14个日期，按照实际日期排序
        # 将MM-DD格式转换为可排序的日期对象
        from datetime import datetime
        current_year = datetime.now().year
        
        def parse_date(date_str):
            try:
                # 处理MM-DD格式
                if '-' in date_str and len(date_str.split('-')) == 2:
                    month, day = date_str.split('-')
                    return datetime(current_year, int(month), int(day))
                # 处理YYYY-MM-DD格式
                elif '-' in date_str and len(date_str.split('-')) == 3:
                    year, month, day = date_str.split('-')
                    return datetime(int(year), int(month), int(day))
                else:
                    return datetime.min
            except:
                return datetime.min
        
        # 按日期排序，最新的在前
        all_dates = sorted(history_data.keys(), key=parse_date, reverse=True)[:14]

        # 创建合并后的数据结构
        merged_data = []
        max_rows = 20  # 每日最多20个代币

        for rank in range(1, max_rows + 1):
            row_data = {'排名': rank}

            # 为每个日期添加数据（14天）
            for i, date in enumerate(all_dates):
                if i == 0:
                    # 第一组列
                    token_key = '代币'
                    date_key = '日期'
                else:
                    # 第2-14组列
                    token_key = f'代币_{i+1}'
                    date_key = f'日期_{i+1}'

                if date in history_data and rank <= len(history_data[date]):
                    token_data = history_data[date][rank-1]
                    row_data[token_key] = token_data['代币']
                    # 确保日期格式为MM-DD（去除年份）
                    if '-' in date and len(date.split('-')) > 2:
                        # 如果日期包含年份，只取月和日
                        date_parts = date.split('-')
                        formatted_date = f"{date_parts[1]}-{date_parts[2]}"
                    else:
                        formatted_date = date
                    row_data[date_key] = formatted_date
                else:
                    row_data[token_key] = ''
                    # 空白位置也要确保日期格式正确
                    if date and '-' in date and len(date.split('-')) > 2:
                        date_parts = date.split('-')
                        formatted_date = f"{date_parts[1]}-{date_parts[2]}"
                    else:
                        formatted_date = date if i < len(all_dates) else ''
                    row_data[date_key] = formatted_date

            merged_data.append(row_data)

        return merged_data

    def save_daily_gainers_history(self, merged_data):
        """保存历史涨幅榜数据 - 14天版本，正确保存所有历史数据"""
        history_file = 'daily_gainers_history.json'

        try:
            # 先加载现有的历史数据
            existing_history = self.load_daily_gainers_history()

            # 从合并数据中提取当前的所有日期数据
            current_history = {}

            # 获取所有日期
            dates = set()
            for row in merged_data:
                for key in row.keys():
                    if key.startswith('日期') and row[key]:
                        dates.add(row[key])

            # 重建历史数据结构
            for date in dates:
                current_history[date] = []

            # 填充每个日期的数据 - 支持14天
            for row in merged_data:
                rank = row['排名']

                # 检查每一列的数据（14天）
                col_suffixes = [''] + [f'_{i}' for i in range(2, 15)]  # '', '_2', '_3', ..., '_14'

                for col_suffix in col_suffixes:
                    date_key = f'日期{col_suffix}' if col_suffix else '日期'
                    token_key = f'代币{col_suffix}' if col_suffix else '代币'

                    if (date_key in row and row[date_key] and
                        token_key in row and row[token_key]):

                        date = row[date_key]
                        token = row[token_key]

                        # 确保这个排名的数据还没有被添加到这个日期
                        existing_ranks = [item['排名'] for item in current_history[date]]
                        if rank not in existing_ranks:
                            current_history[date].append({
                                '排名': rank,
                                '代币': token,
                                '日期': date
                            })

            # 合并现有历史和当前历史
            final_history = existing_history.copy()
            final_history.update(current_history)

            # 移除空的日期条目
            final_history = {date: tokens for date, tokens in final_history.items() if tokens}

            # 对每个日期的数据按排名排序
            for date in final_history:
                final_history[date].sort(key=lambda x: x['排名'])

            with open(history_file, 'w', encoding='utf-8') as f:
                json.dump(final_history, f, ensure_ascii=False, indent=2)
            print(f"💾 历史数据已保存到 {history_file} (包含 {len(final_history)} 个有效日期)")

        except Exception as e:
            print(f"⚠️ 保存历史数据失败: {e}")
            import traceback
            traceback.print_exc()

    def format_gainers_history_for_display(self, merged_data):
        """格式化历史数据用于显示和上传 - 14天版本"""
        df = pd.DataFrame(merged_data)

        # 确保列的顺序正确 - 按照用户要求的格式
        column_order = ['排名']

        # 按照固定顺序添加14组列（代币，日期）
        for i in range(14):  # 固定14天
            if i == 0:
                token_col = '代币'
                date_col = '日期'
            else:
                token_col = f'代币_{i+1}'
                date_col = f'日期_{i+1}'

            # 确保列存在，如果不存在则创建空列
            if token_col not in df.columns:
                df[token_col] = ''
            if date_col not in df.columns:
                df[date_col] = ''

            column_order.extend([token_col, date_col])

        # 重新排序DataFrame
        df = df.reindex(columns=column_order)

        return df

    def get_worksheet_columns(self, dataset_key):
        """Get column configuration for worksheets - Updated per 7.15 requirements"""
        columns_config = {
            'raw_data': [
                ('volume_rank', '排名'),
                ('base_asset', '代币'),
                ('price', '价格(USDT)'),
                ('market_cap', '市值(M USD)'),
                ('total_volume_24h', '24h总量(M USDT)'),
                ('spot_volume_24h', '24h现货交易量(M USDT)'),
                ('futures_volume_24h', '24h合约交易量(M USDT)'),
                ('avg_spot_volume_14d', '14日平均现货交易量(M USDT)'),
                ('avg_futures_volume_14d', '14日期货合约交易量(M USDT)'),
                ('1d_return', '1日涨跌(%)'),
                ('7d_return', '7日涨跌(%)'),
                ('14d_return', '14日涨跌(%)'),
                ('funding_rate', '资金费率(%)'),
                ('open_interest', 'OI持仓量'),
                ('avg_oi_7d', '7日平均OI'),
                ('has_spot', '有现货'),
                ('has_futures', '有期货')
            ],
            'volume_ranking': [
                ('volume_rank', '排名'),
                ('base_asset', '代币'),
                ('price', '价格(USDT)'),
                ('total_volume_24h', '24h总量(M USDT)'),
                ('spot_volume_24h', '24h现货交易量(M USDT)'),
                ('futures_volume_24h', '24h合约交易量(M USDT)'),
                ('avg_spot_volume_14d', '14日现货平均(M USDT)'),
                ('avg_futures_volume_14d', '14日期货平均(M USDT)'),
                ('1d_return', '1日涨跌(%)'),
                ('7d_return', '7日涨跌(%)'),
                ('14d_return', '14日涨跌(%)'),
                ('funding_rate', '资金费率(%)'),
                ('market_cap', '市值(M USD)'),
                ('anomaly_notes', '标注逻辑')
            ],
            'market_cap_ranking': [
                ('mcap_rank', '市值排名'),
                ('base_asset', '代币'),
                ('price', '价格(USDT)'),
                ('market_cap', '市值(M USD)'),
                ('total_volume_24h', '24h总量(M USDT)'),
                ('spot_volume_24h', '24h现货交易量(M USDT)'),
                ('avg_spot_volume_14d', '14日平均现货交易量(M USDT)'),
                ('futures_volume_24h', '24h合约交易量(M USDT)'),
                ('avg_futures_volume_14d', '14日期货合约交易量(M USDT)'),
                ('1d_return', '1日涨跌(%)'),
                ('7d_return', '7日涨跌(%)'),
                ('14d_return', '14日涨跌(%)'),
                ('funding_rate', '资金费率(%)'),
                ('anomaly_notes', '标注逻辑')
            ],
            'gainers_losers': [
                ('gain_loss_rank', '排名'),
                ('gain_loss_type', '类型'),
                ('base_asset', '代币'),
                ('price', '价格(USDT)'),
                ('1d_return', '1日涨跌(%)'),
                ('7d_return', '7日涨跌(%)'),
                ('14d_return', '14日涨跌(%)'),
                ('total_volume_24h', '24h总量(M USDT)'),
                ('spot_volume_24h', '24h现货交易量(M USDT)'),
                ('avg_spot_volume_14d', '14日平均现货交易量(M USDT)'),
                ('futures_volume_24h', '24h合约交易量(M USDT)'),
                ('avg_futures_volume_14d', '14日期货合约交易量(M USDT)'),
                ('funding_rate', '资金费率(%)'),
                ('market_cap', '市值(M USD)'),
                ('anomaly_notes', '标注逻辑')
            ],
            'futures_focus': [
                ('futures_rank', '期货排名'),
                ('base_asset', '代币'),
                ('price', '价格(USDT)'),
                ('futures_volume_24h', '24h合约交易量(M USDT)'),
                ('avg_futures_volume_14d', '14日期货合约交易量(M USDT)'),
                ('open_interest', 'OI持仓量'),
                ('avg_oi_7d', '7日平均OI'),
                ('funding_rate', '资金费率(%)'),
                ('1d_return', '1日涨跌(%)'),
                ('7d_return', '7日涨跌(%)'),
                ('14d_return', '14日涨跌(%)'),
                ('market_cap', '市值(M USD)'),
                ('anomaly_notes', '标注逻辑')
            ],
            'daily_gainers_history': [
                ('排名', '排名'),
                ('代币', '代币'),
                ('日期', '日期'),
                ('代币_2', '代币'),
                ('日期_2', '日期'),
                ('代币_3', '代币'),
                ('日期_3', '日期'),
                ('代币_4', '代币'),
                ('日期_4', '日期'),
                ('代币_5', '代币'),
                ('日期_5', '日期'),
                ('代币_6', '代币'),
                ('日期_6', '日期'),
                ('代币_7', '代币'),
                ('日期_7', '日期'),
                ('代币_8', '代币'),
                ('日期_8', '日期'),
                ('代币_9', '代币'),
                ('日期_9', '日期'),
                ('代币_10', '代币'),
                ('日期_10', '日期'),
                ('代币_11', '代币'),
                ('日期_11', '日期'),
                ('代币_12', '代币'),
                ('日期_12', '日期'),
                ('代币_13', '代币'),
                ('日期_13', '日期'),
                ('代币_14', '代币'),
                ('日期_14', '日期')
            ],
            'low_volume_high_mcap': [
                ('代币', '代币'),
                ('价格(USDT)', '价格(USDT)'),
                ('市值(M USD)', '市值(M USD)'),
                ('24h总量(M USDT)', '24h总量(M USDT)'),
                ('24h现货交易量(M USDT)', '24h现货交易量(M USDT)'),
                ('14日平均现货交易量(M USDT)', '14日平均现货交易量(M USDT)'),
                ('24h合约交易量(M USDT)', '24h合约交易量(M USDT)'),
                ('14日期货合约交易量(M USDT)', '14日期货合约交易量(M USDT)'),
                ('1日涨跌(%)', '1日涨跌(%)'),
                ('7日涨跌', '7日涨跌'),
                ('14日涨跌', '14日涨跌'),
                ('资金费率(%)', '资金费率(%)')
            ],
            'high_volume_low_mcap': [
                ('base_asset', '代币'),
                ('price', '价格(USDT)'),
                ('market_cap', '市值(M USD)'),
                ('mcap_futures_ratio', '市值/合约交易量比值'),
                ('total_volume_24h', '24h总量(M USDT)'),
                ('spot_volume_24h', '24h现货交易量(M USDT)'),
                ('avg_spot_volume_14d', '14日平均现货交易量(M USDT)'),
                ('futures_volume_24h', '24h合约交易量(M USDT)'),
                ('avg_futures_volume_14d', '14日期货合约交易量(M USDT)'),
                ('1d_return', '1日涨跌(%)'),
                ('funding_rate', '资金费率(%)'),
                ('anomaly_notes', '标注逻辑')
            ],
            'low_mcap_high_volume': [
                ('base_asset', '代币'),
                ('price', '价格(USDT)'),
                ('market_cap', '市值(M USD)'),
                ('mcap_futures_ratio', '市值/合约交易量比值'),
                ('total_volume_24h', '24h总量(M USDT)'),
                ('spot_volume_24h', '24h现货交易量(M USDT)'),
                ('avg_spot_volume_14d', '14日平均现货交易量(M USDT)'),
                ('futures_volume_24h', '24h合约交易量(M USDT)'),
                ('avg_futures_volume_14d', '14日期货合约交易量(M USDT)'),
                ('1d_return', '1日涨跌(%)'),
                ('7d_return', '7日涨跌(%)'),
                ('14d_return', '14日涨跌(%)'),
                ('funding_rate', '资金费率(%)'),
                ('anomaly_notes', '标注逻辑')
            ]
        }

        return columns_config.get(dataset_key, [])

    def generate_anomaly_notes(self, df, dataset_key, anomaly_info):
        """Generate anomaly notes for each row - Per 7.15 requirements"""
        notes = []

        for idx, row in df.iterrows():
            note_parts = []

            # Check each anomaly type
            if idx in anomaly_info.get('futures_volume_growth_top5', []):
                if 'futures_volume_24h' in df.columns and 'avg_futures_volume_14d' in df.columns:
                    avg_vol = row['avg_futures_volume_14d']
                    current_vol = row['futures_volume_24h']

                    if avg_vol > 0:  # 避免除零错误
                        ratio = current_vol / avg_vol
                        note_parts.append(f"期货量增长{ratio:.2f}x")
                    else:
                        note_parts.append(f"期货量增长(无历史数据)")

            if idx in anomaly_info.get('spot_volume_growth_top5', []):
                if 'spot_volume_24h' in df.columns and 'avg_spot_volume_14d' in df.columns:
                    avg_vol = row['avg_spot_volume_14d']
                    current_vol = row['spot_volume_24h']

                    if avg_vol > 0:  # 避免除零错误
                        ratio = current_vol / avg_vol
                        note_parts.append(f"现货量增长{ratio:.2f}x")
                    else:
                        note_parts.append(f"现货量增长(无历史数据)")

            if idx in anomaly_info.get('1d_return_top5', []):
                note_parts.append(f"1日涨跌前5")

            if idx in anomaly_info.get('14d_return_top5', []):
                note_parts.append(f"14日涨跌前5")

            if idx in anomaly_info.get('volume_vs_mcap_top5', []):
                note_parts.append(f"量/市值极高")

            if idx in anomaly_info.get('funding_rate_anomaly', []):
                note_parts.append(f"资金费率异常")

            if idx in anomaly_info.get('oi_vs_mcap', []):
                note_parts.append(f"OI/市值高")

            if idx in anomaly_info.get('oi_vs_avg_oi', []):
                note_parts.append(f"OI增长异常")

            # Join all note parts
            final_note = "; ".join(note_parts) if note_parts else ""
            notes.append(final_note)

        return notes

    def create_excel_with_annotations(self, datasets):
        """创建带标注的Excel文件 - 使用pandas简化版本"""
        print("📊 创建带标注的数据文件中...")

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')

        try:
            # 首先尝试创建Excel文件
            excel_filename = f"币安代币分析_{timestamp}.xlsx"
            excel_path = self.get_output_path(excel_filename, 'Excel文件')

            try:
                # 使用pandas ExcelWriter
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:

                    # 为每个数据集创建工作表
                    for dataset_key, df in datasets.items():
                        if dataset_key not in self.worksheets_config:
                            continue

                        worksheet_title = self.worksheets_config[dataset_key]
                        print(f"  📋 处理工作表: {worksheet_title}")

                        # 检测精细化异常
                        anomaly_info = self.detect_refined_anomalies(df, dataset_key)

                        # 生成异常标注
                        if dataset_key not in ['raw_data', 'daily_gainers_history']:
                            anomaly_notes = self.generate_anomaly_notes(df, dataset_key, anomaly_info)
                            df_with_notes = df.copy()
                            df_with_notes['anomaly_notes'] = anomaly_notes
                        else:
                            df_with_notes = df.copy()

                        # 获取列配置
                        columns = self.get_worksheet_columns(dataset_key)
                        if not columns:
                            continue

                        # 格式化数据
                        formatted_df = self.format_data_for_sheets(df_with_notes)

                        # 选择需要的列
                        selected_columns = [col_name for col_name, _ in columns if col_name in formatted_df.columns]
                        display_df = formatted_df[selected_columns].copy()

                        # 重命名列为显示名称
                        column_mapping = {col_name: display_name for col_name, display_name in columns if col_name in display_df.columns}
                        display_df = display_df.rename(columns=column_mapping)

                        # 创建信息头DataFrame
                        timestamp_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        total_anomalies = sum(len(indices) for indices in anomaly_info.values())

                        info_data = {
                            '信息': [
                                f"📅 更新时间: {timestamp_str}",
                                f"🔐 认证方式: OAuth v1.0 (个人账户)",
                                f"📊 数据来源: 币安API (现货:{self.spot_count}, 期货:{self.futures_count})",
                                f"🎯 代币数量: {len(df)} 个代币",
                                f"🔍 精细化异常: {total_anomalies} 个检测到",
                                "",  # 空行
                                "数据开始:"
                            ]
                        }
                        info_df = pd.DataFrame(info_data)

                        # 写入信息头
                        info_df.to_excel(writer, sheet_name=worksheet_title, index=False, header=False)

                        # 写入主数据 (从第8行开始)
                        display_df.to_excel(writer, sheet_name=worksheet_title, startrow=7, index=False)

                        print(f"    ✅ {worksheet_title}: {len(df)} 行数据已处理")

                print(f"✅ Excel文件已保存: {excel_path}")
                return excel_path

            except ImportError:
                print("⚠️ openpyxl未安装，使用CSV格式作为备选方案")
                return self.create_csv_files(datasets, timestamp)

        except Exception as e:
            print(f"❌ 数据文件创建失败: {e}")
            return None

    def create_csv_files(self, datasets, timestamp):
        """创建CSV文件作为备选方案"""
        print("� 创建CSV文件中...")

        csv_files = []

        try:
            # 为每个数据集创建CSV文件
            for dataset_key, df in datasets.items():
                if dataset_key not in self.worksheets_config:
                    continue

                worksheet_title = self.worksheets_config[dataset_key]
                csv_filename = f"币安代币分析_{worksheet_title}_{timestamp}.csv"
                csv_path = self.get_output_path(csv_filename, 'CSV文件')

                print(f"  📋 处理CSV: {worksheet_title}")

                # 检测精细化异常
                anomaly_info = self.detect_refined_anomalies(df, dataset_key)

                # 生成异常标注
                if dataset_key not in ['raw_data', 'daily_gainers_history']:
                    anomaly_notes = self.generate_anomaly_notes(df, dataset_key, anomaly_info)
                    df_with_notes = df.copy()
                    df_with_notes['anomaly_notes'] = anomaly_notes
                else:
                    df_with_notes = df.copy()

                # 获取列配置
                columns = self.get_worksheet_columns(dataset_key)
                if not columns:
                    continue

                # 格式化数据
                formatted_df = self.format_data_for_sheets(df_with_notes)

                # 选择需要的列
                selected_columns = [col_name for col_name, _ in columns if col_name in formatted_df.columns]
                display_df = formatted_df[selected_columns].copy()

                # 重命名列为显示名称
                column_mapping = {col_name: display_name for col_name, display_name in columns if col_name in display_df.columns}
                display_df = display_df.rename(columns=column_mapping)

                # 保存CSV文件
                display_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                csv_files.append(csv_path)

                print(f"    ✅ {csv_filename}: {len(df)} 行数据已保存")

            print(f"✅ 已创建 {len(csv_files)} 个CSV文件")
            return csv_files

        except Exception as e:
            print(f"❌ CSV文件创建失败: {e}")
            return None

    def create_excel_from_csv_files(self, csv_files):
        """从CSV文件创建带颜色标注的Excel文件"""
        print("📊 从CSV文件创建带颜色标注的Excel文件...")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        excel_filename = f"币安代币分析_完整版_{timestamp}.xlsx"
        excel_path = self.get_output_path(excel_filename, 'Excel文件')

        try:
            # 尝试导入openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
                openpyxl_available = True
            except ImportError:
                print("⚠️ openpyxl未安装，使用pandas创建基础Excel文件")
                openpyxl_available = False

            if openpyxl_available:
                return self.create_excel_with_openpyxl(csv_files, excel_path)
            else:
                return self.create_excel_with_pandas(csv_files, excel_path)

        except Exception as e:
            print(f"❌ Excel文件创建失败: {e}")
            return None

    def create_excel_with_pandas(self, csv_files, excel_filename):
        """使用pandas创建基础Excel文件"""
        print("📊 使用pandas创建基础Excel文件...")

        try:
            # 尝试不同的Excel引擎
            try:
                writer = pd.ExcelWriter(excel_filename, engine='openpyxl')
            except ImportError:
                try:
                    writer = pd.ExcelWriter(excel_filename, engine='xlsxwriter')
                except ImportError:
                    # 使用默认引擎
                    writer = pd.ExcelWriter(excel_filename)

            with writer:

                for csv_file in csv_files:
                    # 从文件名提取工作表名称
                    sheet_name = csv_file.split('_')[2].replace('.csv', '')
                    print(f"  📋 处理工作表: {sheet_name}")

                    # 读取CSV文件
                    df = pd.read_csv(csv_file, encoding='utf-8-sig')

                    # 写入Excel工作表
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # 简化版本，不进行复杂格式化
                    try:
                        # 尝试设置列宽（如果支持）
                        if hasattr(writer, 'sheets') and sheet_name in writer.sheets:
                            worksheet = writer.sheets[sheet_name]
                            if hasattr(worksheet, 'set_column'):
                                for i, col in enumerate(df.columns):
                                    max_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                                    worksheet.set_column(i, i, min(max_len, 30))
                    except:
                        # 忽略格式化错误
                        pass

                    print(f"    ✅ {sheet_name}: {len(df)} 行已处理")

            print(f"✅ 基础Excel文件已保存: {excel_filename}")
            return excel_filename

        except Exception as e:
            print(f"❌ pandas Excel创建失败: {e}")
            return None

    def create_excel_with_openpyxl(self, csv_files, excel_filename):
        """使用openpyxl创建带颜色标注的Excel文件"""
        print("📊 使用openpyxl创建带颜色标注的Excel文件...")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment

            wb = Workbook()

            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 定义异常颜色
            colors = {
                'futures_volume_growth': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),  # 浅红色
                'spot_volume_growth': PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid'),     # 浅橙色
                '1d_return': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),             # 浅黄色
                '14d_return': PatternFill(start_color='E5FFCC', end_color='E5FFCC', fill_type='solid'),            # 浅绿色
                'volume_vs_mcap': PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),        # 浅蓝色
                'funding_rate': PatternFill(start_color='FFCCFF', end_color='FFCCFF', fill_type='solid'),          # 浅紫色
                'oi_anomaly': PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid'),            # 浅青色
                'general': PatternFill(start_color='FFCCE5', end_color='FFCCE5', fill_type='solid')                # 浅粉色
            }

            for csv_file in csv_files:
                # 从文件名提取工作表名称
                sheet_name = csv_file.split('_')[2].replace('.csv', '')
                print(f"  📋 处理工作表: {sheet_name}")

                # 读取CSV文件
                df = pd.read_csv(csv_file, encoding='utf-8-sig')

                # 创建工作表
                ws = wb.create_sheet(title=sheet_name)

                # 添加信息头
                timestamp_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                info_data = [
                    [f"📅 更新时间: {timestamp_str}"],
                    [f"🔐 认证方式: OAuth v1.0 (个人账户)"],
                    [f"📊 数据来源: 币安API (现货:{self.spot_count}, 期货:{self.futures_count})"],
                    [f"🎯 代币数量: {len(df)} 个代币"],
                    [f"🔍 精细化异常: 颜色标注显示"],
                    []  # 空行
                ]

                # 写入信息头
                for row_idx, row_data in enumerate(info_data, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        if value:  # 只写入非空值
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.font = Font(bold=True)

                # 写入表头
                header_row = len(info_data) + 1
                for col_idx, header in enumerate(df.columns, 1):
                    cell = ws.cell(row=header_row, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

                # 写入数据并应用颜色标注
                data_start_row = header_row + 1
                for row_idx, (_, row) in enumerate(df.iterrows()):
                    excel_row = data_start_row + row_idx

                    # 写入数据
                    for col_idx, (col_name, value) in enumerate(row.items(), 1):
                        if pd.isna(value):
                            value = ''
                        cell = ws.cell(row=excel_row, column=col_idx, value=str(value))

                        # 应用颜色标注
                        self.apply_excel_color_formatting(cell, col_name, value, row)

                # 调整列宽
                for col_idx, col_name in enumerate(df.columns, 1):
                    max_len = max(df[col_name].astype(str).str.len().max(), len(col_name)) + 2
                    col_letter = chr(64 + col_idx)
                    ws.column_dimensions[col_letter].width = min(max_len, 30)

                print(f"    ✅ {sheet_name}: {len(df)} 行已处理，颜色标注已应用")

            # 保存Excel文件
            wb.save(excel_filename)
            print(f"✅ 带颜色标注的Excel文件已保存: {excel_filename}")

            return excel_filename

        except Exception as e:
            print(f"❌ openpyxl Excel创建失败: {e}")
            # 回退到pandas方式
            return self.create_excel_with_pandas(csv_files, excel_filename.replace('完整版', '基础版'))

    def apply_excel_color_formatting(self, cell, col_name, value, row):
        """应用Excel颜色格式化"""
        try:
            from openpyxl.styles import PatternFill

            # 检查标注逻辑列
            if 'anomaly_notes' in row and pd.notna(row['anomaly_notes']) and row['anomaly_notes']:
                notes = str(row['anomaly_notes'])

                # 根据标注内容应用不同颜色
                if '期货量增长' in notes and col_name in ['24h合约交易量(M USDT)', '14日期货合约交易量(M USDT)']:
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                elif '现货量增长' in notes and col_name in ['24h现货交易量(M USDT)', '14日平均现货交易量(M USDT)']:
                    cell.fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')
                elif '1日涨跌前5' in notes and col_name == '1日涨跌(%)':
                    cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                elif '14日涨跌前5' in notes and col_name == '14日涨跌(%)':
                    cell.fill = PatternFill(start_color='E5FFCC', end_color='E5FFCC', fill_type='solid')
                elif '量/市值极高' in notes and col_name in ['24h总量(M USDT)', '市值(M USD)']:
                    cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                elif '资金费率异常' in notes and col_name == '资金费率(%)':
                    cell.fill = PatternFill(start_color='FFCCFF', end_color='FFCCFF', fill_type='solid')
                elif 'OI' in notes and col_name in ['OI持仓量', '7日平均OI']:
                    cell.fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')

        except Exception as e:
            # 忽略格式化错误，继续处理
            pass

    def upload_files_to_sheets(self, files):
        """将文件上传到Google Sheets"""
        print("📤 将数据文件上传到Google Sheets中...")

        try:
            # 创建或获取Google Sheet
            if not self.create_or_get_sheet():
                print("❌ 创建Google Sheet失败")
                return False

            # 处理不同类型的文件
            if isinstance(files, str):
                # 单个Excel文件
                if files.endswith('.xlsx'):
                    return self.upload_excel_file(files)
                else:
                    print("❌ 不支持的文件格式")
                    return False
            elif isinstance(files, list):
                # 多个CSV文件
                return self.upload_csv_files(files)
            else:
                print("❌ 无效的文件参数")
                return False

        except Exception as e:
            print(f"❌ 文件上传失败: {e}")
            return False

    def upload_excel_file(self, excel_filename):
        """上传Excel文件 - 修复数据格式问题"""
        try:
            # 读取Excel文件
            excel_data = pd.read_excel(excel_filename, sheet_name=None, header=None)

            # 为每个Excel工作表创建Google工作表
            for sheet_name, df in excel_data.items():
                print(f"  📋 上传工作表: {sheet_name}")

                try:
                    # 清理数据
                    cleaned_df = self.clean_data_for_upload(df)

                    # 创建或获取工作表
                    try:
                        worksheet = self.sheet.worksheet(sheet_name)
                        worksheet.clear()
                    except gspread.WorksheetNotFound:
                        worksheet = self.sheet.add_worksheet(title=sheet_name, rows=1000, cols=30)

                    # 准备上传数据
                    upload_data = []
                    for _, row in cleaned_df.iterrows():
                        row_data = []
                        for value in row:
                            # 确保值是字符串且不为空
                            str_value = str(value) if value is not None else ''
                            # 限制字符串长度
                            if len(str_value) > 1000:
                                str_value = str_value[:1000] + '...'
                            row_data.append(str_value)
                        upload_data.append(row_data)

                    # 分批上传数据
                    batch_size = 100
                    for i in range(0, len(upload_data), batch_size):
                        batch_data = upload_data[i:i+batch_size]
                        start_row = i + 1

                        try:
                            worksheet.update(values=batch_data, range_name=f'A{start_row}')
                            print(f"    📤 批次 {i//batch_size + 1}: {len(batch_data)} 行已上传")
                        except Exception as batch_error:
                            print(f"    ⚠️ 批次 {i//batch_size + 1} 上传失败: {batch_error}")
                            continue

                    print(f"    ✅ {sheet_name}: 数据处理完成")

                except Exception as e:
                    print(f"    ❌ {sheet_name} 上传失败: {e}")
                    continue

            print(f"✅ Excel文件已成功上传到Google Sheets")
            print(f"📊 Google Sheet链接: {self.sheet.url}")
            return True

        except Exception as e:
            print(f"❌ Excel上传失败: {e}")
            return False

    def clean_data_for_upload(self, df):
        """清理数据以确保Google Sheets兼容性"""
        cleaned_df = df.copy()

        for col in cleaned_df.columns:
            if cleaned_df[col].dtype in ['float64', 'float32']:
                # 处理无穷大和NaN值
                cleaned_df[col] = cleaned_df[col].replace([float('inf'), float('-inf')], 0)
                cleaned_df[col] = cleaned_df[col].fillna(0)

                # 限制浮点数精度，避免JSON兼容性问题
                cleaned_df[col] = cleaned_df[col].round(6)

                # 将超大数值限制在合理范围内
                max_val = 1e15  # 设置最大值限制
                cleaned_df[col] = cleaned_df[col].clip(-max_val, max_val)

        # 确保所有数据都是字符串格式
        for col in cleaned_df.columns:
            cleaned_df[col] = cleaned_df[col].astype(str)
            # 替换可能导致问题的特殊字符
            cleaned_df[col] = cleaned_df[col].replace(['inf', '-inf', 'nan'], '0')

        return cleaned_df

    def upload_csv_files(self, csv_files):
        """上传CSV文件 - 修复数据格式问题"""
        try:
            for csv_file in csv_files:
                # 从文件名提取工作表名称
                sheet_name = csv_file.split('_')[2].replace('.csv', '')  # 提取工作表名称
                print(f"  📋 上传CSV: {sheet_name}")

                try:
                    # 读取CSV文件
                    df = pd.read_csv(csv_file, encoding='utf-8-sig')

                    # 清理数据
                    cleaned_df = self.clean_data_for_upload(df)

                    # 创建或获取工作表
                    try:
                        worksheet = self.sheet.worksheet(sheet_name)
                        worksheet.clear()
                    except gspread.WorksheetNotFound:
                        worksheet = self.sheet.add_worksheet(title=sheet_name, rows=1000, cols=30)

                    # 添加信息头
                    timestamp_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    info_data = [
                        [f"📅 更新时间: {timestamp_str}"],
                        [f"🔐 认证方式: OAuth v1.0 (个人账户)"],
                        [f"📊 数据来源: 币安API (现货:{self.spot_count}, 期货:{self.futures_count})"],
                        [f"🎯 代币数量: {len(df)} 个代币"],
                        [f"🔍 精细化异常: 单元格级别标注"],
                        [],  # 空行
                        list(cleaned_df.columns)  # 表头
                    ]

                    # 添加数据行 - 确保所有值都是字符串
                    for _, row in cleaned_df.iterrows():
                        row_data = []
                        for value in row:
                            # 确保值是字符串且不为空
                            str_value = str(value) if value is not None else ''
                            # 限制字符串长度，避免过长的数据
                            if len(str_value) > 1000:
                                str_value = str_value[:1000] + '...'
                            row_data.append(str_value)
                        info_data.append(row_data)

                    # 分批上传数据，避免一次性上传过多数据
                    batch_size = 100
                    for i in range(0, len(info_data), batch_size):
                        batch_data = info_data[i:i+batch_size]
                        start_row = i + 1

                        try:
                            worksheet.update(values=batch_data, range_name=f'A{start_row}')
                            print(f"    📤 批次 {i//batch_size + 1}: {len(batch_data)} 行已上传")
                        except Exception as batch_error:
                            print(f"    ⚠️ 批次 {i//batch_size + 1} 上传失败: {batch_error}")
                            continue

                    print(f"    ✅ {sheet_name}: {len(df)} 行数据处理完成")

                except Exception as e:
                    print(f"    ❌ {csv_file} 上传失败: {e}")
                    continue

            print(f"✅ CSV文件已成功上传到Google Sheets")
            print(f"📊 Google Sheet链接: {self.sheet.url}")
            return True

        except Exception as e:
            print(f"❌ CSV上传失败: {e}")
            return False

    def create_or_get_sheet(self):
        """Create or get daily Google Sheet - Updated per 7.15 requirements"""
        try:
            # Check if today's sheet exists
            try:
                self.sheet = self.gc.open(self.sheet_name)
                print(f"📋 Today's sheet exists: {self.sheet_name}")

                # Clear existing worksheets and delete blank Sheet1
                worksheets = self.sheet.worksheets()
                for ws in worksheets:
                    if ws.title.lower() in ['sheet1', 'sheet 1', '工作表1']:
                        print(f"🗑️ Deleting blank sheet: {ws.title}")
                        self.sheet.del_worksheet(ws)
                    else:
                        ws.clear()
                        print(f"🧹 Cleared worksheet: {ws.title}")

            except gspread.SpreadsheetNotFound:
                # Create new sheet
                self.sheet = self.gc.create(self.sheet_name)
                print(f"✅ Created new sheet: {self.sheet_name}")

                # Delete the default Sheet1
                try:
                    default_sheet = self.sheet.sheet1
                    if default_sheet.title.lower() in ['sheet1', 'sheet 1', '工作表1']:
                        print(f"🗑️ Deleting default blank sheet: {default_sheet.title}")
                        self.sheet.del_worksheet(default_sheet)
                except:
                    pass

            # Set permissions - 修改为所有人可编辑
            try:
                self.sheet.share(self.user_email, perm_type='user', role='writer')
                print(f"✅ Granted edit access to {self.user_email}")
            except:
                print("⚠️ User permission setting failed")

            try:
                # 修改为所有人都可以编辑
                self.sheet.share('', perm_type='anyone', role='writer')
                print("✅ Set as publicly editable - 所有人可编辑")
            except:
                print("⚠️ Public permission setting failed")

            print(f"📊 Today's sheet URL: {self.sheet.url}")
            return True

        except Exception as e:
            print(f"❌ Sheet operation failed: {e}")
            return False

    def apply_refined_anomaly_formatting(self, worksheet, df, anomaly_info, columns):
        """Apply refined single-cell anomaly formatting - Per 7.15 requirements"""
        try:
            # Define colors for different anomaly types
            colors = {
                'futures_volume_growth_top5': {'red': 1.0, 'green': 0.8, 'blue': 0.8},  # Light red
                'spot_volume_growth_top5': {'red': 1.0, 'green': 0.9, 'blue': 0.8},     # Light orange
                '1d_return_top5': {'red': 1.0, 'green': 1.0, 'blue': 0.8},             # Light yellow
                '14d_return_top5': {'red': 0.9, 'green': 1.0, 'blue': 0.8},            # Light green
                'volume_vs_mcap_top5': {'red': 0.8, 'green': 0.9, 'blue': 1.0},        # Light blue
                'funding_rate_anomaly': {'red': 1.0, 'green': 0.8, 'blue': 1.0},       # Light purple
                'oi_vs_mcap': {'red': 0.8, 'green': 1.0, 'blue': 1.0},                 # Light cyan
                'oi_vs_avg_oi': {'red': 1.0, 'green': 0.8, 'blue': 0.9}                # Light pink
            }

            # Create column index mapping
            col_mapping = {}
            for i, (col_name, _) in enumerate(columns):
                col_mapping[col_name] = chr(65 + i)  # A, B, C, etc.

            # Apply formatting for each anomaly type
            for anomaly_type, indices in anomaly_info.items():
                if not indices:
                    continue

                color = colors.get(anomaly_type, {'red': 1.0, 'green': 0.8, 'blue': 0.8})

                # Determine which columns to highlight based on anomaly type
                target_columns = []
                if 'futures_volume' in anomaly_type:
                    target_columns = ['futures_volume_24h', 'avg_futures_volume_14d']
                elif 'spot_volume' in anomaly_type:
                    target_columns = ['spot_volume_24h', 'avg_spot_volume_14d']
                elif '1d_return' in anomaly_type:
                    target_columns = ['1d_return']
                elif '14d_return' in anomaly_type:
                    target_columns = ['14d_return']
                elif 'volume_vs_mcap' in anomaly_type:
                    target_columns = ['total_volume_24h', 'market_cap']
                elif 'funding_rate' in anomaly_type:
                    target_columns = ['funding_rate']
                elif 'oi_vs_mcap' in anomaly_type:
                    target_columns = ['open_interest', 'market_cap']
                elif 'oi_vs_avg_oi' in anomaly_type:
                    target_columns = ['open_interest', 'avg_oi_7d']

                # Apply formatting to specific cells
                for row_idx in indices:
                    actual_row = row_idx + 7  # Add header offset
                    for col_name in target_columns:
                        if col_name in col_mapping:
                            col_letter = col_mapping[col_name]
                            range_name = f"{col_letter}{actual_row}"

                            worksheet.format(range_name, {
                                "backgroundColor": color,
                                "textFormat": {"bold": True}
                            })

            total_formatted = sum(len(indices) for indices in anomaly_info.values())
            print(f"  🎨 Applied refined anomaly highlighting: {total_formatted} cells")

        except Exception as e:
            print(f"  ⚠️ Refined formatting failed: {e}")

    # 旧的工作表创建函数已被Excel生成方式替代，减少API调用

    def run_analysis(self):
        """运行完整代币分析 - 主执行方法"""
        self.display_version_info()

        # 获取用户偏好
        self.get_user_preferences()

        # 询问是否使用离线模式
        offline_mode = self.check_offline_mode()

        if not offline_mode:
            # OAuth认证
            if not self.authenticate_oauth():
                print("❌ OAuth认证失败")
                print("🔄 切换到离线模式...")
                offline_mode = True

        # 获取增强市场数据
        enhanced_df = self.get_enhanced_market_data()
        if enhanced_df is None:
            print("❌ 获取增强市场数据失败")
            return None

        # 创建增强数据集
        datasets = self.create_enhanced_datasets(enhanced_df)

        # 创建带标注的数据文件
        data_files = self.create_excel_with_annotations(datasets)
        if not data_files:
            print("❌ 数据文件创建失败")
            return None

        # 创建整合的Excel文件
        if isinstance(data_files, list):  # CSV文件列表
            print("📊 整合CSV文件到Excel...")
            excel_file = self.create_excel_from_csv_files(data_files)
            if excel_file:
                print(f"✅ 整合Excel文件已创建: {excel_file}")
                final_files = {'csv': data_files, 'excel': excel_file}
            else:
                final_files = data_files
        else:  # 已经是Excel文件
            final_files = data_files

        if offline_mode:
            print("📁 离线模式: 仅生成本地文件")
            self.display_results(datasets, enhanced_df, final_files, offline=True)
            return datasets
        else:
            # 询问是否上传到Google Sheets（自动模式下默认上传）
            if self.auto_mode:
                upload_choice = 'y'
                print("\n🤖 自动模式：默认上传到Google Sheets")
            else:
                upload_choice = input("\n🔗 是否上传到Google Sheets? (y/n, 默认n): ").strip().lower()

            if upload_choice == 'y':
                # 将数据文件上传到Google Sheets
                upload_file = excel_file if isinstance(final_files, dict) and 'excel' in final_files else data_files
                upload_success = self.upload_files_to_sheets(upload_file)

                if upload_success:
                    self.display_results(datasets, enhanced_df, final_files)
                    return datasets
                else:
                    print("\n❌ 数据上传失败，但本地文件已生成")
                    self.display_results(datasets, enhanced_df, final_files, offline=True)
                    return datasets
            else:
                print("📁 跳过Google Sheets上传，仅保留本地文件")
                self.display_results(datasets, enhanced_df, final_files, offline=True)
                return datasets

    def check_offline_mode(self):
        """检查是否使用离线模式"""
        # 自动模式下默认使用在线模式
        if self.auto_mode:
            print("\n🤖 自动模式：使用在线模式")
            return False  # 在线模式
            
        print("\n🌐 网络模式选择:")
        print("1. 在线模式 (生成本地文件 + 上传到Google Sheets)")
        print("2. 离线模式 (仅生成本地文件)")

        while True:
            choice = input("请选择模式 (1/2, 默认1): ").strip()
            if not choice:
                choice = "1"

            if choice == "1":
                return False  # 在线模式
            elif choice == "2":
                return True   # 离线模式
            else:
                print("❌ 请输入1或2")
                continue

    def display_results(self, datasets, enhanced_df, data_files=None, offline=False):
        """显示分析结果"""
        print(f"\n🎉 币安代币筛选器 v{self.version} - 分析完成!")
        print(f"📊 成功创建 {len(datasets)} 个工作表")
        print(f"📁 输出文件夹: {self.output_folder}")

        if offline:
            print("📁 离线模式: 本地文件已生成")

        if data_files:
            if isinstance(data_files, dict):
                # 包含CSV和Excel文件
                if 'csv' in data_files:
                    csv_files = data_files['csv']
                    print(f"📁 本地CSV文件: {len(csv_files)} 个文件")
                if 'excel' in data_files:
                    excel_file = data_files['excel']
                    print(f"📊 整合Excel文件: {excel_file}")
            elif isinstance(data_files, str):
                print(f"📁 本地Excel文件: {data_files}")
            elif isinstance(data_files, list):
                print(f"📁 本地CSV文件: {len(data_files)} 个文件")
                for file in data_files[:3]:  # 显示前3个文件名
                    print(f"   - {file}")
                if len(data_files) > 3:
                    print(f"   - ... 还有 {len(data_files) - 3} 个文件")

        if not offline and hasattr(self, 'sheet') and self.sheet:
            print(f"🔗 Google Sheets链接: {self.sheet.url}")
        elif offline:
            print("💡 离线模式提示: 如需上传到Google Sheets，请稍后在网络正常时重新运行")

        # 显示统计信息
        total_tokens = len(enhanced_df)

        print(f"\n📈 分析统计:")
        print(f"  - 分析代币总数: {total_tokens}")
        print(f"  - 请求现货代币: {self.spot_count}")
        print(f"  - 请求期货代币: {self.futures_count}")
        print(f"  - 精细化异常检测: 单元格级别标注 (7.15需求)")
        print(f"  - 认证方式: OAuth v{self.version} (个人账户)")

        # 显示工作表详情
        print(f"\n📋 工作表详情:")
        for key, df in datasets.items():
            if key in self.worksheets_config:
                print(f"  - {self.worksheets_config[key]}: {len(df)} 行")

        print(f"\n🆕 v{self.version} 功能 (根据7.15需求更新):")
        print(f"  ✅ 用户可配置代币数量 (现货:{self.spot_count}, 期货:{self.futures_count})")
        print(f"  ✅ 14日历史数据分析")
        print(f"  ✅ 实时资金费率和持仓量数据")
        print(f"  ✅ 精细化单元格异常标注")
        print(f"  ✅ 排除代币过滤 (ALPACA, BNX等)")
        print(f"  ✅ 每日独立文件 (不覆盖历史)")
        print(f"  ✅ OAuth认证 (解决403存储错误)")
        print(f"  ✅ 异常标注逻辑列")
        print(f"  ✅ 删除空白Sheet1标签页")
        print(f"  ✅ 新增缺失标签页: 每日涨幅榜, 低量高市值, 高市值低期货量")
        print(f"  ✅ Excel本地生成 + Google Sheets上传 (减少API调用)")

    def debug_futures_focus_completeness(self, df, futures_symbols):
        """
        Debug function to identify missing futures entries
        用于识别期货专注中缺失条目的调试功能
        """
        print("🔍 调试期货专注完整性...")
        
        # 识别应该在期货专注中的代币
        expected_futures_tokens = set()
        for symbol in futures_symbols:
            if symbol.endswith('USDT'):
                base_asset = symbol.replace('USDT', '')
                # 处理特殊代币映射
                if symbol.startswith('1000'):
                    base_asset = symbol.replace('1000', '').replace('USDT', '')
                expected_futures_tokens.add(base_asset)
        
        actual_futures_tokens = set(df[df['has_futures'] == True]['base_asset'])
        
        missing_tokens = expected_futures_tokens - actual_futures_tokens
        
        if missing_tokens:
            print(f"⚠️  期货专注缺失代币: {list(missing_tokens)[:10]}")  # 只显示前10个
            
            # 检查是否因为没有现货交易对而缺失
            missing_no_spot = []
            for token in missing_tokens:
                spot_symbol = f"{token}USDT"
                if spot_symbol not in df['symbol'].values:
                    missing_no_spot.append(token)
            
            if missing_no_spot:
                print(f"📊 因无现货交易对缺失的代币: {missing_no_spot[:5]}")
        else:
            print("✅ 期货专注无缺失代币")
        
        print(f"📊 期货合约总数: {len(expected_futures_tokens)}")
        print(f"📊 期货专注现有代币数: {len(actual_futures_tokens)}")
        
        return missing_tokens, expected_futures_tokens

    def create_futures_only_placeholder(self, base_asset, futures_symbol, futures_data):
        """
        Create placeholder row for futures-only tokens
        为仅有期货合约的代币创建占位行
        """
        try:
            # 从期货数据中获取基本信息
            volume_data = futures_data.get(futures_symbol, {})
            
            placeholder = {
                'symbol': f"{base_asset}USDT",  # 标准化符号格式
                'base_asset': base_asset,
                'price': volume_data.get('price', 0),
                'has_futures': True,
                'futures_volume_24h': float(volume_data.get('quoteVolume', 0)) if volume_data.get('quoteVolume') else 0,
                'funding_rate': 0,  # 将在后续步骤中填充
                'open_interest': 0,  # 将在后续步骤中填充
                'avg_oi_7d': 0,
                # 现货数据设为默认值
                'volume': 0,
                'spot_volume_24h': 0,
                'market_cap': 0,
                'total_volume_24h': float(volume_data.get('quoteVolume', 0)) if volume_data.get('quoteVolume') else 0,
                '1d_return': float(volume_data.get('priceChangePercent', 0)) if volume_data.get('priceChangePercent') else 0,
                '7d_return': 0,
                '14d_return': 0,
                'avg_spot_volume_14d': 0,
                'avg_futures_volume_14d': 0,
                'count': float(volume_data.get('count', 0)) if volume_data.get('count') else 0,
                'high': float(volume_data.get('highPrice', 0)) if volume_data.get('highPrice') else 0,
                'low': float(volume_data.get('lowPrice', 0)) if volume_data.get('lowPrice') else 0,
                'open': float(volume_data.get('openPrice', 0)) if volume_data.get('openPrice') else 0
            }
            
            return placeholder
        except Exception as e:
            print(f"⚠️  创建占位符失败 {base_asset}: {e}")
            return None

    def create_comprehensive_futures_focus(self, df, futures_symbols, futures_market_data):
        """
        Enhanced futures focus creation ensuring top N futures are included
        增强的期货专注创建，确保前N个期货代币被收录
        """
        print(f"🔍 创建期货专注数据集 (前{self.futures_count}个期货)...")
        
        # 获取交易量最高的前N个期货
        sorted_futures = sorted(
            futures_market_data.items(),
            key=lambda x: x[1]['quoteVolume'],  # 按USD交易额排序
            reverse=True
        )[:self.futures_count]  # 只取前N个
        
        # 提取前N个期货的符号
        top_futures_symbols = {item[0] for item in sorted_futures}
        top_futures_base_assets = set()
        for symbol in top_futures_symbols:
            base_asset = symbol.replace('USDT', '')
            if symbol.startswith('1000'):
                base_asset = symbol.replace('1000', '').replace('USDT', '')
            top_futures_base_assets.add(base_asset)
        
        # 从现有数据中获取期货数据
        futures_df = df[df['has_futures'] == True].copy()
        
        # 过滤只保留前N个期货
        futures_df = futures_df[futures_df['base_asset'].isin(top_futures_base_assets)].copy()
        initial_count = len(futures_df)
        
        # 检查哪些前N个期货缺失
        existing_base_assets = set(futures_df['base_asset'].values)
        missing_top_futures = top_futures_base_assets - existing_base_assets
        
        # 仅为缺失的前N个期货添加占位数据
        added_count = 0
        if missing_top_futures and futures_market_data:
            print(f"🔄 添加 {len(missing_top_futures)} 个缺失的TOP期货代币...")
            
            for futures_symbol, data in sorted_futures:
                base_asset = futures_symbol.replace('USDT', '')
                if futures_symbol.startswith('1000'):
                    base_asset = futures_symbol.replace('1000', '').replace('USDT', '')
                
                if base_asset in missing_top_futures:
                    placeholder_row = self.create_futures_only_placeholder(
                        base_asset, futures_symbol, futures_market_data
                    )
                    if placeholder_row:
                        new_row_df = pd.DataFrame([placeholder_row])
                        futures_df = pd.concat([futures_df, new_row_df], ignore_index=True)
                        added_count += 1
                        if added_count <= 5:  # 只显示前5个
                            print(f"  ✅ 已添加TOP期货: {base_asset} (交易额: ${data['quoteVolume']/1000000:.1f}M)")
        
        # 按期货交易量排序并限制为前N个
        futures_df = futures_df.sort_values('futures_volume_24h', ascending=False).reset_index(drop=True)
        futures_df = futures_df.head(self.futures_count)  # 确保只返回前N个
        futures_df['futures_rank'] = range(1, len(futures_df) + 1)
        
        final_count = len(futures_df)
        print(f"📊 期货专注数据集:")
        print(f"  - 配置的期货数量: {self.futures_count}")
        print(f"  - 初始匹配代币数: {initial_count}")
        print(f"  - 添加的缺失TOP期货: {added_count}")
        print(f"  - 最终期货数量: {final_count}")
        
        return futures_df

def main():
    """主函数 - v2.0入口点"""
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='币安代币筛选器 v2.0')
    parser.add_argument('--auto', action='store_true', 
                        help='自动模式运行（跳过所有用户输入，使用默认值）')
    parser.add_argument('--spot-count', type=int, default=80,
                        help='现货代币数量（默认: 80）')
    parser.add_argument('--futures-count', type=int, default=80,
                        help='期货代币数量（默认: 80）')
    args = parser.parse_args()

    print("🚀 启动币安代币筛选器 v2.0")
    print("生产版本，支持OAuth认证和CoinGecko集成")
    print("=" * 80)
    
    if args.auto:
        print("🤖 自动模式已启用")

    try:
        screener = BinanceTokenScreenerV1(auto_mode=args.auto)
        
        # 如果提供了命令行参数，覆盖默认值
        if args.spot_count:
            screener.spot_count = args.spot_count
        if args.futures_count:
            screener.futures_count = args.futures_count
            
        result = screener.run_analysis()

        if result:
            print("\n✅ 分析完成成功!")
            print("💡 v1.0 生产功能:")
            print("  - OAuth认证 (解决403错误)")
            print("  - 用户可配置代币数量")
            print("  - 14日历史数据分析")
            print("  - 实时资金费率")
            print("  - 智能异常检测")
            print("  - 颜色编码异常标注")
            print("  - 每日独立谷歌表格")
        else:
            print("\n❌ 分析失败")
            print("💡 故障排除:")
            print("  1. 检查OAuth认证状态")
            print("  2. 验证网络连接")
            print("  3. 确认谷歌账户权限")

    except KeyboardInterrupt:
        print("\n⚠️ 用户中断分析")
    except Exception as e:
        print(f"\n❌ 分析错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
